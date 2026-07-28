"""Napari widget for ROI-driven neuron navigation and label visibility."""

from __future__ import annotations

from contextlib import suppress
from pathlib import Path

import napari
import numpy as np
from napari.layers import Labels, Vectors
from napari.utils import colormaps as cmap
from qtpy.QtCore import Qt
from qtpy.QtGui import QBrush, QCloseEvent, QColor, QFont
from qtpy.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSlider,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ._roi import (
    NeuronBoxDataset,
    add_time_axis,
    box_vectors_2d,
    box_vectors_3d,
    neuron_id_to_label_value,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


ROLE_KEY = "napari_label_manager.role"
ROLE_SELECTED = "roi_boxes_selected"
LEGACY_ROLE_ALL = "roi_boxes_all"
ROLE_ACTIVE = "roi_box_active"
MANAGED_VECTOR_ROLES = (ROLE_SELECTED, LEGACY_ROLE_ALL, ROLE_ACTIVE)
MAX_EXACT_LABEL_PIXELS = 10_000_000


class LabelManager(QWidget):
    """Manage Labels visibility and navigate read-only neuron boxes."""

    def __init__(self, napari_viewer: napari.Viewer, parent=None):
        super().__init__(parent)
        self.viewer = napari_viewer
        self.current_layer: Labels | None = None
        self.roi_dataset: NeuronBoxDataset | None = None
        self.active_id: int | None = None
        self.checked_ids: set[int] = set()
        self._available_ids: list[int] = []
        self._selection_items: dict[int, QTreeWidgetItem] = {}
        self._base_colors: dict[int, tuple[float, float, float, float]] = {}
        self._original_colormap = None
        self._original_opacity: float | None = None
        self._ui_sync = False
        self._closed = False
        self._keys_bound: list[str] = []

        self._setup_ui()
        self._connect_viewer_events()
        self._bind_keys()
        self._refresh_label_layers()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _setup_ui(self) -> None:
        layout = QVBoxLayout()
        self.setMinimumWidth(430)

        header = QLabel("Neuron ROI Navigator")
        header.setFont(QFont("Arial", 12, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        self.labels_layer_group = self._build_layer_group()
        layout.addWidget(self.labels_layer_group)
        layout.addWidget(self._build_roi_group())
        layout.addWidget(self._build_selection_group())
        layout.addWidget(self._build_annotation_group())
        layout.addWidget(self._build_status_group())
        self.setLayout(layout)

    def _build_layer_group(self) -> QGroupBox:
        group = QGroupBox("Labels Layer")
        group_layout = QVBoxLayout()
        self.layer_combo = QComboBox()
        self.layer_combo.currentTextChanged.connect(self._on_layer_changed)
        group_layout.addWidget(QLabel("Controlled layer:"))
        group_layout.addWidget(self.layer_combo)
        self._add_labels_appearance_controls(group_layout)
        group.setLayout(group_layout)
        return group

    def _build_roi_group(self) -> QGroupBox:
        group = QGroupBox("Neuron ROI Source")
        group_layout = QVBoxLayout()

        path_layout = QHBoxLayout()
        self.roi_path_input = QLineEdit()
        self.roi_path_input.setReadOnly(True)
        self.roi_path_input.setPlaceholderText("Load neuron_pt_tuple.npy")
        self.load_roi_btn = QPushButton("Load NPY")
        self.load_roi_btn.clicked.connect(self.load_roi_npy)
        self.unload_roi_btn = QPushButton("Unload")
        self.unload_roi_btn.clicked.connect(self.unload_roi)
        self.unload_roi_btn.setEnabled(False)
        path_layout.addWidget(self.roi_path_input, 1)
        path_layout.addWidget(self.load_roi_btn)
        path_layout.addWidget(self.unload_roi_btn)
        group_layout.addLayout(path_layout)

        config_layout = QHBoxLayout()
        config_layout.addWidget(QLabel("Z divisor:"))
        self.z_divisor_spin = QDoubleSpinBox()
        self.z_divisor_spin.setDecimals(3)
        self.z_divisor_spin.setRange(0.001, 1_000_000)
        self.z_divisor_spin.setValue(5.0)
        config_layout.addWidget(self.z_divisor_spin)

        config_layout.addWidget(QLabel("Volume start:"))
        self.volume_start_spin = QSpinBox()
        self.volume_start_spin.setRange(0, 1_000_000)
        config_layout.addWidget(self.volume_start_spin)

        config_layout.addWidget(QLabel("Stride:"))
        self.volume_stride_spin = QSpinBox()
        self.volume_stride_spin.setRange(1, 1_000_000)
        self.volume_stride_spin.setValue(1)
        config_layout.addWidget(self.volume_stride_spin)
        group_layout.addLayout(config_layout)

        self.roi_info_label = QLabel("No ROI loaded; IDs come from Labels")
        self.roi_info_label.setWordWrap(True)
        group_layout.addWidget(self.roi_info_label)
        group.setLayout(group_layout)
        return group

    def _build_selection_group(self) -> QGroupBox:
        group = QGroupBox("Neuron Selection")
        group_layout = QVBoxLayout()

        self.navigation_help_label = QLabel("Q/W: last/next")
        group_layout.addWidget(self.navigation_help_label)

        controls_layout = QHBoxLayout()
        self.check_all_btn = QPushButton("All")
        self.check_all_btn.clicked.connect(self.check_all)
        self.check_none_btn = QPushButton("None")
        self.check_none_btn.clicked.connect(self.check_none)
        controls_layout.addWidget(self.check_all_btn)
        controls_layout.addWidget(self.check_none_btn)
        controls_layout.addStretch(1)
        group_layout.addLayout(controls_layout)

        self.selection_tree = QTreeWidget()
        self.selection_tree.setColumnCount(2)
        self.selection_tree.setHeaderLabels(["", "Neuron"])
        self.selection_tree.setRootIsDecorated(False)
        self.selection_tree.setAlternatingRowColors(False)
        self.selection_tree.setSelectionMode(QAbstractItemView.SingleSelection)
        self.selection_tree.setMinimumHeight(190)
        self.selection_tree.header().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.selection_tree.header().setSectionResizeMode(
            1, QHeaderView.Stretch
        )
        self.selection_tree.itemChanged.connect(
            self._on_selection_item_changed
        )
        self.selection_tree.itemClicked.connect(
            self._on_selection_item_clicked
        )
        group_layout.addWidget(self.selection_tree)
        group.setLayout(group_layout)
        return group

    def _add_labels_appearance_controls(
        self, group_layout: QVBoxLayout
    ) -> None:
        selected_layout = QHBoxLayout()
        selected_layout.addWidget(QLabel("Checked label opacity:"))
        self.selected_opacity_slider = QSlider(Qt.Horizontal)
        self.selected_opacity_slider.setRange(0, 100)
        self.selected_opacity_slider.setValue(100)
        self.selected_opacity_label = QLabel("1.00")
        self.selected_opacity_slider.valueChanged.connect(
            self._on_opacity_changed
        )
        selected_layout.addWidget(self.selected_opacity_slider)
        selected_layout.addWidget(self.selected_opacity_label)
        group_layout.addLayout(selected_layout)

        other_layout = QHBoxLayout()
        other_layout.addWidget(QLabel("Unchecked label opacity:"))
        self.other_opacity_slider = QSlider(Qt.Horizontal)
        self.other_opacity_slider.setRange(0, 100)
        self.other_opacity_slider.setValue(20)
        self.other_opacity_label = QLabel("0.20")
        self.other_opacity_slider.valueChanged.connect(
            self._on_opacity_changed
        )
        other_layout.addWidget(self.other_opacity_slider)
        other_layout.addWidget(self.other_opacity_label)
        group_layout.addLayout(other_layout)

        controls_layout = QHBoxLayout()
        self.hide_unchecked_checkbox = QCheckBox("Hide unchecked labels")
        self.hide_unchecked_checkbox.toggled.connect(
            self._on_opacity_changed
        )
        controls_layout.addWidget(self.hide_unchecked_checkbox)
        controls_layout.addStretch(1)
        group_layout.addLayout(controls_layout)

    def _build_annotation_group(self) -> QGroupBox:
        group = QGroupBox("Neuron Annotation")
        group_layout = QVBoxLayout()
        controls = QHBoxLayout()

        controls.addWidget(QLabel("Sheet:"))
        self.sheet_name_input = QLineEdit("Neuron Annotations")
        self.sheet_name_input.setFixedWidth(130)
        controls.addWidget(self.sheet_name_input)

        controls.addWidget(QLabel("Start:"))
        self.annotation_start_input = QLineEdit("0")
        self.annotation_start_input.setFixedWidth(45)
        controls.addWidget(self.annotation_start_input)
        controls.addWidget(QLabel("End:"))
        self.annotation_end_input = QLineEdit("9")
        self.annotation_end_input.setFixedWidth(45)
        controls.addWidget(self.annotation_end_input)

        self.fill_annotation_btn = QPushButton("Fill")
        self.fill_annotation_btn.clicked.connect(self.fill_annotation_range)
        controls.addWidget(self.fill_annotation_btn)
        self.load_current_labels_btn = QPushButton("Current IDs")
        self.load_current_labels_btn.clicked.connect(
            self.load_current_ids_to_annotation
        )
        controls.addWidget(self.load_current_labels_btn)

        self.load_excel_btn = QPushButton("Load")
        self.load_excel_btn.clicked.connect(self.load_excel_to_annotation)
        self.load_excel_btn.setEnabled(EXCEL_AVAILABLE)
        controls.addWidget(self.load_excel_btn)
        self.save_annotation_btn = QPushButton("Save")
        self.save_annotation_btn.clicked.connect(self.save_annotation_to_excel)
        self.save_annotation_btn.setEnabled(EXCEL_AVAILABLE)
        controls.addWidget(self.save_annotation_btn)
        group_layout.addLayout(controls)

        self.annotation_table = QTableWidget(0, 3)
        self.annotation_table.setHorizontalHeaderLabels(
            ["digital", "biological", "annotation"]
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.Stretch
        )
        self.annotation_table.setSelectionBehavior(
            QAbstractItemView.SelectRows
        )
        self.annotation_table.setSelectionMode(
            QAbstractItemView.SingleSelection
        )
        self.annotation_table.itemSelectionChanged.connect(
            self._on_annotation_selection_changed
        )
        self.annotation_table.itemChanged.connect(
            self._on_annotation_item_changed
        )
        group_layout.addWidget(self.annotation_table)
        group.setLayout(group_layout)
        return group

    def _build_status_group(self) -> QGroupBox:
        group = QGroupBox("Status")
        group_layout = QVBoxLayout()
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: green;")
        group_layout.addWidget(self.status_label)
        self.info_text = QTextEdit()
        self.info_text.setReadOnly(True)
        self.info_text.setMaximumHeight(90)
        group_layout.addWidget(self.info_text)
        group.setLayout(group_layout)
        return group

    # ------------------------------------------------------------------
    # Viewer and layer lifecycle
    # ------------------------------------------------------------------
    def _connect_viewer_events(self) -> None:
        events = self.viewer.layers.events
        events.inserted.connect(self._refresh_label_layers)
        events.removed.connect(self._on_layer_removed)
        events.reordered.connect(self._refresh_label_layers)
        if hasattr(events, "renamed"):
            events.renamed.connect(self._refresh_label_layers)

        dims_events = self.viewer.dims.events
        dims_events.point.connect(self._on_dims_changed)
        dims_events.ndisplay.connect(self._on_dims_changed)
        dims_events.order.connect(self._on_dims_changed)

    def _disconnect_viewer_events(self) -> None:
        events = self.viewer.layers.events
        for emitter, callback in (
            (events.inserted, self._refresh_label_layers),
            (events.removed, self._on_layer_removed),
            (events.reordered, self._refresh_label_layers),
        ):
            with suppress(TypeError, ValueError):
                emitter.disconnect(callback)
        if hasattr(events, "renamed"):
            with suppress(TypeError, ValueError):
                events.renamed.disconnect(self._refresh_label_layers)

        dims_events = self.viewer.dims.events
        for emitter in (
            dims_events.point,
            dims_events.ndisplay,
            dims_events.order,
        ):
            with suppress(TypeError, ValueError):
                emitter.disconnect(self._on_dims_changed)

    def _bind_keys(self) -> None:
        bindings = (("Q", self._previous_key), ("W", self._next_key))
        for key, callback in bindings:
            try:
                self.viewer.bind_key(key, callback)
                self._keys_bound.append(key)
            except ValueError:
                self.update_status(
                    f"Hotkey {key} is already assigned; button navigation remains available",
                    "orange",
                )

    def _unbind_keys(self) -> None:
        for key in self._keys_bound:
            with suppress(KeyError, ValueError):
                self.viewer.bind_key(key, None, overwrite=True)
        self._keys_bound.clear()

    def _refresh_label_layers(self, event=None) -> None:
        del event
        if self._closed:
            return
        current_name = (
            self.current_layer.name if self.current_layer is not None else ""
        )
        names = [
            layer.name
            for layer in self.viewer.layers
            if isinstance(layer, Labels)
        ]

        self.layer_combo.blockSignals(True)
        self.layer_combo.clear()
        if names:
            self.layer_combo.addItems(names)
            if current_name in names:
                self.layer_combo.setCurrentText(current_name)
            self.layer_combo.setEnabled(True)
        else:
            self.layer_combo.addItem("No Labels layers available")
            self.layer_combo.setEnabled(False)
        self.layer_combo.blockSignals(False)

        target_name = self.layer_combo.currentText() if names else ""
        if target_name != current_name or self.current_layer is None:
            self._on_layer_changed(target_name)

    def _on_layer_changed(self, layer_name: str) -> None:
        if self.current_layer is not None and self.current_layer.name == layer_name:
            return

        self._detach_current_layer(restore=True)
        self._remove_vector_layers()
        if not layer_name or layer_name == "No Labels layers available":
            self.current_layer = None
            self._available_ids = []
            self.checked_ids.clear()
            self.active_id = None
            self._rebuild_selection_items()
            self._remove_vector_layers()
            self.update_status("No Labels layer selected", "orange")
            return

        try:
            layer = self.viewer.layers[layer_name]
        except KeyError:
            self.update_status("Labels layer no longer exists", "red")
            return
        if not isinstance(layer, Labels):
            self.update_status("Selected layer is not a Labels layer", "red")
            return

        self.current_layer = layer
        self._original_colormap = layer.colormap
        self._original_opacity = float(layer.opacity)
        self._base_colors.clear()
        layer.events.data.connect(self._on_label_data_changed)
        self.checked_ids.clear()
        self.active_id = None
        self._refresh_available_ids(select_first=True)
        if self.roi_dataset is not None and layer.ndim in (3, 4):
            self._ensure_vector_layers()
        elif self.roi_dataset is not None:
            self.update_status(
                "ROI overlays require a (z,y,x) or (t,z,y,x) Labels layer",
                "orange",
            )
        self._refresh_vector_layers()
        self.update_status(f"Selected layer: {layer.name}", "blue")

    def _detach_current_layer(self, *, restore: bool) -> None:
        if self.current_layer is None:
            return
        with suppress(TypeError, ValueError):
            self.current_layer.events.data.disconnect(
                self._on_label_data_changed
            )
        if restore:
            self._restore_layer_display()
        self.current_layer = None
        self._original_colormap = None
        self._original_opacity = None
        self._base_colors.clear()

    def _on_layer_removed(self, event) -> None:
        if self._closed:
            return
        removed = getattr(event, "value", None)
        if (
            isinstance(removed, Vectors)
            and removed.metadata.get(ROLE_KEY) in MANAGED_VECTOR_ROLES
        ):
            return
        if removed is self.current_layer:
            self._detach_current_layer(restore=False)
        self._refresh_label_layers()

    def _on_label_data_changed(self, event=None) -> None:
        del event
        if self.roi_dataset is None:
            self._refresh_available_ids(select_first=False)
        else:
            self._cache_base_colors(self._available_ids)
            self._apply_label_opacity()
        self._refresh_vector_layers()

    def closeEvent(self, event: QCloseEvent) -> None:
        self.shutdown()
        super().closeEvent(event)

    def shutdown(self) -> None:
        """Restore managed display state and disconnect all external events."""
        if self._closed:
            return
        self._closed = True
        self._disconnect_viewer_events()
        self._unbind_keys()
        self._detach_current_layer(restore=True)
        self._remove_vector_layers()

    # ------------------------------------------------------------------
    # ID discovery, checkable selection, and navigation
    # ------------------------------------------------------------------
    def _refresh_available_ids(self, *, select_first: bool) -> None:
        if self.roi_dataset is not None:
            ids = self.roi_dataset.neuron_ids
        else:
            ids = self._exact_label_ids()
        self._available_ids = ids

        if self.active_id not in ids:
            self.active_id = None
        self.checked_ids.intersection_update(ids)

        if select_first and ids and self.active_id is None:
            valid = self._valid_navigation_ids()
            self.active_id = valid[0] if valid else ids[0]
            self.checked_ids = {self.active_id}

        self._cache_base_colors(ids)
        self._rebuild_selection_items()
        self._apply_label_opacity()

    def _exact_label_ids(self) -> list[int]:
        if self.current_layer is None:
            return []
        data = self.current_layer.data
        size = getattr(data, "size", None)
        if not isinstance(data, np.ndarray) or size is None:
            self.info_text.setText(
                "Out-of-core Labels are not scanned automatically. "
                "Load neuron_pt_tuple.npy for authoritative IDs."
            )
            return []
        if int(size) > MAX_EXACT_LABEL_PIXELS:
            self.info_text.setText(
                f"Labels has {int(size):,} voxels. Automatic ID discovery "
                "is disabled above 10,000,000 voxels; load ROI NPY."
            )
            return []

        unique_values = np.unique(data)
        background = self._background_value()
        return sorted(
            int(value)
            for value in unique_values
            if int(value) != background
        )

    def _background_value(self) -> int:
        colormap = (
            self._original_colormap
            if self._original_colormap is not None
            else getattr(self.current_layer, "colormap", None)
        )
        return int(getattr(colormap, "background_value", 0))

    def _label_value(self, display_id: int) -> int:
        if self.roi_dataset is not None:
            return neuron_id_to_label_value(display_id)
        return int(display_id)

    def _cache_base_colors(self, display_ids: list[int]) -> None:
        if self.current_layer is None or self._original_colormap is None:
            return
        label_values = {self._label_value(item_id) for item_id in display_ids}
        data = self.current_layer.data
        size = getattr(data, "size", None)
        if (
            isinstance(data, np.ndarray)
            and size is not None
            and int(size) <= MAX_EXACT_LABEL_PIXELS
        ):
            background = self._background_value()
            label_values.update(
                int(value)
                for value in np.unique(data)
                if int(value) != background
            )

        managed_colormap = self.current_layer.colormap
        if managed_colormap is not self._original_colormap:
            self.current_layer.colormap = self._original_colormap
        try:
            for label_value in label_values:
                if label_value in self._base_colors:
                    continue
                mapped = np.asarray(
                    self.current_layer.get_color(label_value), dtype=float
                ).reshape(-1)
                if mapped.size != 4 or not np.all(np.isfinite(mapped)):
                    raise ValueError(
                        "Labels colormap returned an invalid RGBA value "
                        f"for label {label_value}"
                    )
                self._base_colors[label_value] = tuple(mapped)
        finally:
            if managed_colormap is not self._original_colormap:
                self.current_layer.colormap = managed_colormap

    def _rebuild_selection_items(self) -> None:
        self._ui_sync = True
        try:
            self.selection_tree.clear()
            self._selection_items.clear()
            for display_id in self._available_ids:
                item = QTreeWidgetItem()
                item.setFlags(
                    item.flags()
                    | Qt.ItemIsUserCheckable
                    | Qt.ItemIsSelectable
                    | Qt.ItemIsEnabled
                )
                item.setData(0, Qt.UserRole, display_id)
                item.setCheckState(
                    0,
                    Qt.Checked
                    if display_id in self.checked_ids
                    else Qt.Unchecked,
                )
                self.selection_tree.addTopLevelItem(item)
                self._selection_items[display_id] = item
        finally:
            self._ui_sync = False
        self._refresh_selection_item_styles()

    def _refresh_selection_item_styles(self) -> None:
        names = self._annotation_names()
        valid_ids = set(self._valid_navigation_ids())
        self._ui_sync = True
        try:
            for display_id in self._available_ids:
                self._style_selection_item(
                    display_id,
                    valid=display_id in valid_ids,
                    biological_name=names.get(display_id, ""),
                )
            if self.active_id is None:
                self.selection_tree.clearSelection()
                self.selection_tree.setCurrentItem(None)
            else:
                item = self._selection_items.get(self.active_id)
                if item is not None:
                    self.selection_tree.setCurrentItem(item)
                    self.selection_tree.scrollToItem(item)
        finally:
            self._ui_sync = False

    def _style_selection_item(
        self,
        display_id: int,
        *,
        valid: bool | None = None,
        biological_name: str = "",
    ) -> None:
        item = self._selection_items.get(display_id)
        if item is None:
            return
        if valid is None:
            valid = display_id in set(self._valid_navigation_ids())
        if not biological_name:
            biological_name = self._annotation_names().get(display_id, "")

        prefix = "Neuron" if self.roi_dataset is not None else "Label"
        text = f"{prefix} {display_id}"
        if biological_name:
            text += f" · {biological_name}"
        if self.roi_dataset is not None and not valid:
            text += " (missing)"
        item.setText(1, text)
        item.setCheckState(
            0,
            Qt.Checked if display_id in self.checked_ids else Qt.Unchecked,
        )
        font = item.font(1)
        font.setBold(display_id == self.active_id)
        item.setFont(1, font)
        if valid:
            rgba = self._display_color(display_id)
            color = QColor.fromRgbF(*rgba[:3])
        else:
            color = QColor("#777777")
        item.setForeground(1, QBrush(color))

    def _display_color(self, display_id: int) -> np.ndarray:
        label_value = self._label_value(display_id)
        rgba = self._base_colors.get(label_value, (0.7, 0.7, 0.7, 1.0))
        return np.asarray(rgba, dtype=float)

    def _on_selection_item_changed(
        self, item: QTreeWidgetItem, column: int
    ) -> None:
        if self._ui_sync or column != 0:
            return
        display_id = int(item.data(0, Qt.UserRole))
        checked = item.checkState(0) == Qt.Checked
        if checked:
            self.checked_ids.add(display_id)
            self.active_id = display_id
        else:
            self.checked_ids.discard(display_id)
            if self.active_id == display_id:
                self.active_id = None
        self._selection_changed(locate=checked)

    def _on_selection_item_clicked(
        self, item: QTreeWidgetItem, column: int
    ) -> None:
        if self._ui_sync or column == 0:
            return
        display_id = int(item.data(0, Qt.UserRole))
        self.activate_id(display_id, locate=True)

    def activate_id(self, display_id: int, *, locate: bool = True) -> None:
        """Check and activate one available identity."""
        if display_id not in self._available_ids:
            return
        self.checked_ids.add(display_id)
        self.active_id = display_id
        self._selection_changed(locate=locate)

    def check_all(self) -> None:
        """Check every available identity without changing an existing active."""
        self.checked_ids = set(self._available_ids)
        locate = False
        if self.active_id is None:
            valid = self._valid_navigation_ids()
            if valid:
                self.active_id = valid[0]
                locate = True
        self._selection_changed(locate=locate)

    def check_none(self) -> None:
        """Clear both the checked collection and active identity."""
        self.checked_ids.clear()
        self.active_id = None
        self._selection_changed(locate=False)

    def _selection_changed(self, *, locate: bool) -> None:
        self._refresh_selection_item_styles()
        self._apply_label_opacity()
        self._refresh_vector_layers()
        self._sync_annotation_to_active()
        if locate:
            self._locate_active_box()
        self._update_info()

    def _valid_navigation_ids(self) -> list[int]:
        if self.roi_dataset is None:
            return list(self._available_ids)
        return self.roi_dataset.valid_ids(self._viewer_time())

    def navigate(self, step: int) -> None:
        valid_ids = self._valid_navigation_ids()
        if not valid_ids:
            self.update_status("No valid neuron at the current volume", "orange")
            return
        if self.active_id not in valid_ids:
            next_id = valid_ids[0] if step >= 0 else valid_ids[-1]
        else:
            index = valid_ids.index(self.active_id)
            next_id = valid_ids[(index + step) % len(valid_ids)]
        self.activate_id(next_id, locate=True)

    def _previous_key(self, viewer=None) -> None:
        del viewer
        self.navigate(-1)

    def _next_key(self, viewer=None) -> None:
        del viewer
        self.navigate(1)

    # ------------------------------------------------------------------
    # Labels appearance
    # ------------------------------------------------------------------
    def _on_opacity_changed(self, value=None) -> None:
        del value
        self.selected_opacity_label.setText(
            f"{self.selected_opacity_slider.value() / 100:.2f}"
        )
        self.other_opacity_slider.setEnabled(
            not self.hide_unchecked_checkbox.isChecked()
        )
        other = (
            0.0
            if self.hide_unchecked_checkbox.isChecked()
            else self.other_opacity_slider.value() / 100
        )
        self.other_opacity_label.setText(f"{other:.2f}")
        self._apply_label_opacity()

    def _apply_label_opacity(self) -> None:
        if (
            self.current_layer is None
            or self._original_colormap is None
            or not self._available_ids
        ):
            return

        selected_alpha = self.selected_opacity_slider.value() / 100
        other_alpha = (
            0.0
            if self.hide_unchecked_checkbox.isChecked()
            else self.other_opacity_slider.value() / 100
        )
        colors: dict[int | None, tuple[float, float, float, float]] = {
            None: (0.0, 0.0, 0.0, 0.0)
        }
        for label_value, base in self._base_colors.items():
            display_id = (
                label_value - 1
                if self.roi_dataset is not None
                else label_value
            )
            rgba = list(base)
            rgba[3] = (
                selected_alpha
                if display_id in self.checked_ids
                else other_alpha
            )
            colors[label_value] = tuple(rgba)

        direct = cmap.direct_colormap(colors)
        direct.background_value = self._background_value()
        direct.name = "neuron_roi_visibility"
        self.current_layer.colormap = direct
        self.current_layer.opacity = 1.0

    def _restore_layer_display(self) -> None:
        if self.current_layer is None:
            return
        if self._original_colormap is not None:
            self.current_layer.colormap = self._original_colormap
        if self._original_opacity is not None:
            self.current_layer.opacity = self._original_opacity

    # ------------------------------------------------------------------
    # ROI loading and geometry layers
    # ------------------------------------------------------------------
    def load_roi_npy(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Load neuron_pt_tuple",
            "",
            "NumPy arrays (*.npy);;All files (*)",
        )
        if not path:
            return

        try:
            self.load_roi_path(path)
        except (OSError, RuntimeError, TypeError, ValueError) as error:
            QMessageBox.critical(
                self, "Invalid ROI file", f"Could not load ROI data:\n{error}"
            )
            self.update_status("ROI load failed", "red")

    def load_roi_path(self, path: str | Path) -> None:
        """Load and activate a read-only ROI NPY without opening a dialog."""
        if self.current_layer is None:
            raise RuntimeError("Select a Labels layer before loading ROI data")
        if self.current_layer.ndim not in (3, 4):
            raise ValueError(
                "ROI overlays require a (z,y,x) or (t,z,y,x) Labels layer"
            )

        dataset = NeuronBoxDataset.from_npy(
            path,
            z_divisor=self.z_divisor_spin.value(),
            volume_start=self.volume_start_spin.value(),
            volume_stride=self.volume_stride_spin.value(),
        )
        source_path = Path(path)

        self.roi_dataset = dataset
        self.roi_path_input.setText(str(source_path))
        self.unload_roi_btn.setEnabled(True)
        self._set_roi_config_enabled(False)
        self.active_id = None
        self.checked_ids.clear()
        self._refresh_available_ids(select_first=True)
        self._ensure_vector_layers()
        self._refresh_vector_layers()
        self._update_roi_info()
        self.update_status(f"Loaded ROI: {source_path.name}", "green")

    def unload_roi(self) -> None:
        self.roi_dataset = None
        self.roi_path_input.clear()
        self.unload_roi_btn.setEnabled(False)
        self._set_roi_config_enabled(True)
        self._remove_vector_layers()
        self.active_id = None
        self.checked_ids.clear()
        self.roi_info_label.setText("No ROI loaded; IDs come from Labels")
        self._refresh_available_ids(select_first=True)
        self.update_status("ROI unloaded", "green")

    def _set_roi_config_enabled(self, enabled: bool) -> None:
        self.z_divisor_spin.setEnabled(enabled)
        self.volume_start_spin.setEnabled(enabled)
        self.volume_stride_spin.setEnabled(enabled)
        self.load_roi_btn.setEnabled(enabled)

    def _update_roi_info(self) -> None:
        if self.roi_dataset is None:
            return
        viewer_t = self._viewer_time()
        source_t = self.roi_dataset.source_time(viewer_t)
        valid = len(self.roi_dataset.valid_ids(viewer_t))
        self.roi_info_label.setText(
            f"T={self.roi_dataset.time_count}, "
            f"N={self.roi_dataset.neuron_count}; "
            f"viewer t={viewer_t} → source t={source_t}; "
            f"{valid} valid boxes"
        )

    def _viewer_time(self) -> int:
        if self.current_layer is None or self.current_layer.ndim == 3:
            return 0
        steps = self.viewer.dims.current_step
        return int(steps[-4]) if len(steps) >= 4 else 0

    def _viewer_z(self) -> int:
        if self.current_layer is None:
            return 0
        steps = self.viewer.dims.current_step
        if self.current_layer.ndim == 4 and len(steps) >= 4:
            return int(steps[-3])
        if len(steps) >= 3:
            return int(steps[-3])
        return 0

    def _shape_zyx(self) -> tuple[int, int, int]:
        if self.current_layer is None:
            raise RuntimeError("No Labels layer selected")
        shape = self.current_layer.data.shape
        return tuple(int(value) for value in shape[-3:])

    def _view_axes_supported(self) -> bool:
        if self.current_layer is None:
            return False
        ndim = int(self.viewer.dims.ndim)
        order = tuple(self.viewer.dims.order)
        if self.viewer.dims.ndisplay == 2:
            return order[-2:] == (ndim - 2, ndim - 1)
        return order[-3:] == (ndim - 3, ndim - 2, ndim - 1)

    def _on_dims_changed(self, event=None) -> None:
        del event
        if (
            self.roi_dataset is not None
            and self.current_layer is not None
            and self.current_layer in self.viewer.layers
        ):
            self._update_roi_info()
            self._refresh_selection_item_styles()
            self._refresh_vector_layers()

    def _ensure_vector_layers(self) -> None:
        if (
            self.current_layer is None
            or self.roi_dataset is None
            or self.current_layer.ndim not in (3, 4)
        ):
            return
        ndim = self.current_layer.ndim
        empty = np.empty((0, 2, ndim), dtype=float)
        axis_labels = tuple(self.current_layer.axis_labels)

        legacy_layer = self._managed_vector_layer(LEGACY_ROLE_ALL)
        if legacy_layer is not None and legacy_layer in self.viewer.layers:
            self.viewer.layers.remove(legacy_layer)

        if self._managed_vector_layer(ROLE_SELECTED) is None:
            self.viewer.add_vectors(
                empty,
                ndim=ndim,
                name="Neuron boxes – selected",
                vector_style="line",
                edge_width=1.0,
                edge_color="white",
                opacity=0.25,
                blending="translucent",
                scale=tuple(self.current_layer.scale),
                translate=tuple(self.current_layer.translate),
                axis_labels=axis_labels,
                metadata={ROLE_KEY: ROLE_SELECTED},
            )
        if self._managed_vector_layer(ROLE_ACTIVE) is None:
            self.viewer.add_vectors(
                empty,
                ndim=ndim,
                name="Neuron box – active",
                vector_style="line",
                edge_width=3.0,
                edge_color="yellow",
                opacity=1.0,
                blending="translucent",
                scale=tuple(self.current_layer.scale),
                translate=tuple(self.current_layer.translate),
                axis_labels=axis_labels,
                metadata={ROLE_KEY: ROLE_ACTIVE},
            )

    def _managed_vector_layer(self, role: str) -> Vectors | None:
        for layer in self.viewer.layers:
            if (
                isinstance(layer, Vectors)
                and layer.metadata.get(ROLE_KEY) == role
            ):
                return layer
        return None

    def _remove_vector_layers(self) -> None:
        managed = [
            layer
            for layer in self.viewer.layers
            if isinstance(layer, Vectors)
            and layer.metadata.get(ROLE_KEY) in MANAGED_VECTOR_ROLES
        ]
        for layer in managed:
            if layer in self.viewer.layers:
                self.viewer.layers.remove(layer)

    def _refresh_vector_layers(self) -> None:
        if (
            self.roi_dataset is None
            or self.current_layer is None
            or self.current_layer.ndim not in (3, 4)
        ):
            return
        selected_layer = self._managed_vector_layer(ROLE_SELECTED)
        active_layer = self._managed_vector_layer(ROLE_ACTIVE)
        if selected_layer is None or active_layer is None:
            return

        if not self._view_axes_supported():
            self._set_vector_data(selected_layer, [], [])
            self._set_vector_data(active_layer, [], [], active=True)
            self.update_status(
                "ROI overlays require spatial y/x axes in 2D or z/y/x axes in 3D",
                "orange",
            )
            return

        viewer_t = self._viewer_time()
        z_index = self._viewer_z()
        shape_zyx = self._shape_zyx()
        selected_vectors: list[np.ndarray] = []
        selected_neuron_ids: list[int] = []
        active_vectors: list[np.ndarray] = []
        active_ids: list[int] = []

        for neuron_id in self.roi_dataset.valid_ids(viewer_t):
            if (
                neuron_id not in self.checked_ids
                and neuron_id != self.active_id
            ):
                continue
            box = self.roi_dataset.get_box(viewer_t, neuron_id)
            if box is None:
                continue
            if self.viewer.dims.ndisplay == 2:
                geometry = box_vectors_2d(
                    box, z_index, shape_zyx=shape_zyx
                )
            else:
                geometry = box_vectors_3d(box, shape_zyx=shape_zyx)
            if self.current_layer.ndim == 4:
                geometry = add_time_axis(geometry, viewer_t)
            if len(geometry):
                if neuron_id in self.checked_ids:
                    selected_vectors.append(geometry)
                    selected_neuron_ids.extend(
                        [neuron_id] * len(geometry)
                    )
                if neuron_id == self.active_id:
                    active_vectors.append(geometry)
                    active_ids.extend([neuron_id] * len(geometry))

        self._set_vector_data(
            selected_layer, selected_vectors, selected_neuron_ids
        )
        self._set_vector_data(
            active_layer, active_vectors, active_ids, active=True
        )

    def _set_vector_data(
        self,
        layer: Vectors,
        vector_blocks: list[np.ndarray],
        neuron_ids: list[int],
        *,
        active: bool = False,
    ) -> None:
        ndim = layer.ndim
        data = (
            np.concatenate(vector_blocks, axis=0)
            if vector_blocks
            else np.empty((0, 2, ndim), dtype=float)
        )
        layer.data = data
        layer.features = {"neuron_id": np.asarray(neuron_ids, dtype=int)}
        if active:
            layer.edge_color = "yellow"
        elif neuron_ids:
            layer.edge_color = np.asarray(
                [self._display_color(item_id) for item_id in neuron_ids]
            )
        else:
            layer.edge_color = "white"

    def _locate_active_box(self) -> None:
        if (
            self.roi_dataset is None
            or self.current_layer is None
            or self.active_id is None
        ):
            return
        box = self.roi_dataset.get_box(self._viewer_time(), self.active_id)
        if box is None:
            self.update_status(
                f"Neuron {self.active_id} is missing at this volume", "orange"
            )
            return

        z, y, x = box.center_zyx
        steps = list(self.viewer.dims.current_step)
        if self.current_layer.ndim == 4 and len(steps) >= 4:
            z_axis = len(steps) - 3
        else:
            z_axis = len(steps) - 3
        z_size = self._shape_zyx()[0]
        steps[z_axis] = int(np.clip(round(z), 0, z_size - 1))
        self.viewer.dims.current_step = tuple(steps)

        point = (
            (self._viewer_time(), z, y, x)
            if self.current_layer.ndim == 4
            else (z, y, x)
        )
        world = self.current_layer.data_to_world(point)
        self.viewer.camera.center = tuple(world[-3:])

    # ------------------------------------------------------------------
    # Annotation and Excel
    # ------------------------------------------------------------------
    def _annotation_names(self) -> dict[int, str]:
        names: dict[int, str] = {}
        for row in range(self.annotation_table.rowCount()):
            digital = self.annotation_table.item(row, 0)
            biological = self.annotation_table.item(row, 1)
            if digital is None:
                continue
            try:
                item_id = int(digital.text())
            except ValueError:
                continue
            names[item_id] = biological.text() if biological is not None else ""
        return names

    def _annotation_rows(
        self,
    ) -> dict[int, tuple[str, str]]:
        rows: dict[int, tuple[str, str]] = {}
        for row in range(self.annotation_table.rowCount()):
            digital = self.annotation_table.item(row, 0)
            if digital is None:
                continue
            try:
                item_id = int(digital.text())
            except ValueError:
                continue
            biological = self.annotation_table.item(row, 1)
            annotation = self.annotation_table.item(row, 2)
            rows[item_id] = (
                biological.text() if biological is not None else "",
                annotation.text() if annotation is not None else "",
            )
        return rows

    def _set_annotation_rows(
        self,
        rows: list[tuple[int, str, str]],
    ) -> None:
        self._ui_sync = True
        self.annotation_table.blockSignals(True)
        try:
            self.annotation_table.setRowCount(len(rows))
            for row, (item_id, biological, annotation) in enumerate(rows):
                digital_item = QTableWidgetItem(str(item_id))
                digital_item.setTextAlignment(Qt.AlignCenter)
                self.annotation_table.setItem(row, 0, digital_item)
                self.annotation_table.setItem(
                    row, 1, QTableWidgetItem(biological)
                )
                self.annotation_table.setItem(
                    row, 2, QTableWidgetItem(annotation)
                )
        finally:
            self.annotation_table.blockSignals(False)
            self._ui_sync = False
        self._refresh_selection_item_styles()

    def fill_annotation_range(self) -> None:
        try:
            start = int(self.annotation_start_input.text())
            end = int(self.annotation_end_input.text())
        except ValueError:
            QMessageBox.warning(
                self, "Invalid range", "Start and end must be integers."
            )
            return
        if start > end:
            QMessageBox.warning(
                self, "Invalid range", "Start cannot be greater than end."
            )
            return
        existing = self._annotation_rows()
        rows = [
            (item_id, *existing.get(item_id, ("", "")))
            for item_id in range(start, end + 1)
        ]
        self._set_annotation_rows(rows)

    def load_current_ids_to_annotation(self) -> None:
        existing = self._annotation_rows()
        rows = [
            (item_id, *existing.get(item_id, ("", "")))
            for item_id in self._available_ids
        ]
        self._set_annotation_rows(rows)
        self.update_status(
            f"Loaded {len(rows)} identities into annotation table", "green"
        )

    def _sync_annotation_to_active(self) -> None:
        if self.active_id is None:
            self._ui_sync = True
            try:
                self.annotation_table.clearSelection()
            finally:
                self._ui_sync = False
            return
        self._ui_sync = True
        try:
            for row in range(self.annotation_table.rowCount()):
                item = self.annotation_table.item(row, 0)
                if item is not None and item.text() == str(self.active_id):
                    self.annotation_table.selectRow(row)
                    self.annotation_table.scrollToItem(item)
                    break
        finally:
            self._ui_sync = False

    def _on_annotation_selection_changed(self) -> None:
        if self._ui_sync:
            return
        rows = self.annotation_table.selectionModel().selectedRows()
        if not rows:
            return
        item = self.annotation_table.item(rows[0].row(), 0)
        if item is None:
            return
        try:
            item_id = int(item.text())
        except ValueError:
            return
        if item_id not in self._available_ids:
            return
        self.activate_id(item_id, locate=True)

    def _on_annotation_item_changed(self, item: QTableWidgetItem) -> None:
        if self._ui_sync:
            return
        if item.column() in (0, 1):
            self._refresh_selection_item_styles()

    def save_annotation_to_excel(self) -> None:
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(
                self, "Excel unavailable", "Install the 'excel' extra first."
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save neuron annotations",
            "neuron_annotations.xlsx",
            "Excel workbooks (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        sheet_name = self.sheet_name_input.text().strip() or "Neuron Annotations"

        try:
            if Path(path).exists():
                workbook = load_workbook(path)
                if sheet_name in workbook.sheetnames:
                    reply = QMessageBox.question(
                        self,
                        "Sheet exists",
                        f"Overwrite sheet '{sheet_name}'?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No,
                    )
                    if reply != QMessageBox.Yes:
                        workbook.close()
                        return
                    workbook.remove(workbook[sheet_name])
                sheet = workbook.create_sheet(sheet_name)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = sheet_name

            sheet.append(["digital", "biological", "annotation"])
            for cell in sheet[1]:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            for item_id, (biological, annotation) in self._annotation_rows().items():
                sheet.append([item_id, biological, annotation])
            for column in sheet.columns:
                width = max(
                    (len(str(cell.value)) for cell in column if cell.value is not None),
                    default=0,
                )
                sheet.column_dimensions[column[0].column_letter].width = (
                    width + 2
                )
            workbook.save(path)
            workbook.close()
        except (OSError, PermissionError, ValueError) as error:
            QMessageBox.critical(
                self, "Save failed", f"Could not save workbook:\n{error}"
            )
            self.update_status("Annotation save failed", "red")
            return
        self.update_status(f"Saved annotations to {sheet_name}", "green")

    def load_excel_to_annotation(self) -> None:
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(
                self, "Excel unavailable", "Install the 'excel' extra first."
            )
            return
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Load neuron annotations",
            "",
            "Excel workbooks (*.xlsx)",
        )
        if not path:
            return
        requested = self.sheet_name_input.text().strip()

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            if requested and requested in workbook.sheetnames:
                sheet = workbook[requested]
            else:
                sheet = workbook[workbook.sheetnames[0]]
                self.sheet_name_input.setText(sheet.title)

            rows: list[tuple[int, str, str]] = []
            for values in sheet.iter_rows(values_only=True):
                if not values or values[0] is None:
                    continue
                try:
                    numeric_id = float(values[0])
                except (TypeError, ValueError):
                    continue
                if (
                    not np.isfinite(numeric_id)
                    or not numeric_id.is_integer()
                    or numeric_id < 0
                ):
                    continue
                item_id = int(numeric_id)
                biological = (
                    "" if len(values) < 2 or values[1] is None else str(values[1])
                )
                annotation = (
                    "" if len(values) < 3 or values[2] is None else str(values[2])
                )
                rows.append((item_id, biological, annotation))
            workbook.close()
        except (OSError, PermissionError, ValueError) as error:
            QMessageBox.critical(
                self, "Load failed", f"Could not load workbook:\n{error}"
            )
            self.update_status("Annotation load failed", "red")
            return

        self._set_annotation_rows(rows)
        self.update_status(f"Loaded {len(rows)} annotations", "green")

    # ------------------------------------------------------------------
    # Status helpers
    # ------------------------------------------------------------------
    def _update_info(self) -> None:
        mode = "ROI" if self.roi_dataset is not None else "Labels"
        checked = sorted(self.checked_ids)
        text = (
            f"Mode: {mode}\n"
            f"Active: {self.active_id if self.active_id is not None else 'none'}\n"
            f"Checked: {checked[:12]}"
        )
        if len(checked) > 12:
            text += f" … (+{len(checked) - 12})"
        self.info_text.setText(text)

    def update_status(self, message: str, color: str = "black") -> None:
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {color};")
