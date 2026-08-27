"""Napari widget for ROI-driven neuron navigation and annotation."""

from __future__ import annotations

from contextlib import suppress
from pathlib import Path

import napari
import numpy as np
from napari.layers import Image, Points, Vectors
from qtpy.QtCore import QEvent, QObject, Qt
from qtpy.QtGui import QBrush, QCloseEvent, QColor, QFont, QPalette
from qtpy.QtWidgets import (
    QAbstractItemView,
    QAbstractSpinBox,
    QApplication,
    QCheckBox,
    QColorDialog,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ._colors import neuron_color
from ._orientation import (
    ALLOWED_ROTATIONS,
    Orientation2D,
    OrientationState,
    resolve_orientation,
)
from ._roi import (
    NeuronBoxDataset,
    add_time_axis,
    box_label_point_2d,
    box_label_point_3d,
    box_vectors_2d,
    box_vectors_3d,
)
from ._z_layers import (
    ZLayerRange,
    build_z_layer_ranges,
    find_z_layer,
    parse_z_cuts,
    shifted_z_translation,
    slice_z_range,
    z_threshold_count_profile,
)
from ._z_profile import ZThresholdProfileWidget

try:
    # The model is optional while older installations are being upgraded.  The
    # widget remains usable for browse-only ROI navigation when it is absent.
    from ._proofread import ProofreadStore
except ImportError:  # pragma: no cover - exercised by compatibility installs
    ProofreadStore = None  # type: ignore[assignment,misc]

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# Preserve the legacy metadata namespace so managed layers from earlier
# releases remain identifiable during cleanup.
ROLE_KEY = "napari_label_manager.role"
ROLE_SELECTED = "roi_boxes_selected"
LEGACY_ROLE_ALL = "roi_boxes_all"
ROLE_ACTIVE = "roi_box_active"
ROLE_BOX_LABELS = "roi_box_labels"
ROLE_PROOF_TARGET = "roi_proof_target"
# The proofreading target is a display-only crosshair.  Keep its geometry
# independent from the ROI box dimensions so users can tune visual length and
# line weight without changing annotation data.
PROOF_TARGET_HALF_LENGTH = 8.0
PROOF_TARGET_EDGE_WIDTH = 1.0
MANAGED_VECTOR_ROLES = (ROLE_SELECTED, LEGACY_ROLE_ALL, ROLE_ACTIVE)
MANAGED_ROI_ROLES = (
    *MANAGED_VECTOR_ROLES,
    ROLE_BOX_LABELS,
    ROLE_PROOF_TARGET,
)
ROLE_Z_IMAGE = "z_layer_image"
MANAGED_Z_ROLES = (ROLE_Z_IMAGE,)
LABEL_MODE_BIOLOGICAL = "biological"
LABEL_MODE_DIGITAL = "digital"
LABEL_MODE_DIGITAL_BIOLOGICAL = "digital_biological"
Z_SOURCE_GEOMETRY_EVENTS = (
    "scale",
    "translate",
    "rotate",
    "shear",
    "affine",
    "axis_labels",
    "units",
)


def _match_neuron_ids(
    available_ids: list[int],
    biological_names: dict[int, str],
    query: str,
) -> list[int]:
    """Return global IDs matching comma-separated digital/name tokens."""
    tokens = [token.strip() for token in query.split(",") if token.strip()]
    if not tokens:
        return []

    digital_ids: set[int] = set()
    biological_tokens: list[str] = []
    for token in tokens:
        if token.isdecimal():
            with suppress(ValueError):
                digital_ids.add(int(token))
        else:
            biological_tokens.append(token.casefold())
    matches: list[int] = []
    for neuron_id in available_ids:
        biological = biological_names.get(neuron_id, "").strip().casefold()
        if neuron_id in digital_ids or any(
            token in biological for token in biological_tokens
        ):
            matches.append(neuron_id)
    return matches


class _ProofreadingKeyFilter(QObject):
    """Temporarily consume proofreading function keys on the canvas.

    napari's layer keymap owns several ordinary keys.  Installing this small
    filter only while proofreading is enabled keeps those bindings intact and
    lets Qt editors receive keys normally.
    """

    _KEY_ACTIONS = {
        Qt.Key_F7: "_proof_delete_current",
        Qt.Key_F8: "_proof_place_cursor",
        Qt.Key_F9: "_proof_add_neuron",
        Qt.Key_F12: "_proof_cancel_or_exit",
    }

    def __init__(self, widget: NeuronAnnotatorWidget) -> None:
        super().__init__(widget)
        self.widget = widget

    def eventFilter(self, watched, event):  # noqa: N802 - Qt API
        del watched
        if event.type() != QEvent.KeyPress:
            return False
        key = int(event.key())
        action = self._KEY_ACTIONS.get(key)
        if action is None or event.modifiers() != Qt.NoModifier:
            return False
        if self.widget._proof_key_focus_blocked():
            return False
        if not self.widget.proofreading_enabled:
            return False
        getattr(self.widget, action)()
        event.accept()
        return True


class NeuronAnnotatorWidget(QWidget):
    """Navigate read-only neuron boxes on an Image."""

    def __init__(self, napari_viewer: napari.Viewer, parent=None):
        super().__init__(parent)
        self.viewer = napari_viewer
        self.current_image: Image | None = None
        self.roi_dataset: NeuronBoxDataset | None = None
        self.active_id: int | None = None
        self.checked_ids: set[int] = set()
        self._available_ids: list[int] = []
        self._selection_items: dict[int, QTreeWidgetItem] = {}
        self._search_match_ids: list[int] = []
        self._search_cursor = -1
        self._box_label_color = "#ffffff"
        self._ui_sync = False
        self._closed = False
        self._keys_bound: list[str] = []
        self._last_viewer_time: int | None = None
        self._z_ranges: tuple[ZLayerRange, ...] = ()
        self._z_source_image: Image | None = None
        self._z_image_layers: list[Image] = []
        self._z_active_index: int | None = None
        self._z_source_image_visible: bool | None = None
        self._z_cleanup = False
        self._z_session_token = f"{id(self):x}"
        self._z_profile_source: Image | None = None
        self._z_profile_time: int | None = None
        self._orientation_state = OrientationState()
        self._orientation_baseline_order: tuple[int, ...] | None = None
        self._orientation_baseline_camera: Orientation2D | None = None
        self._orientation_baseline_ndim: int | None = None
        self._orientation_applying = False
        self._orientation_ui_sync = False
        self._proof_detached = False
        # Proofreading state is deliberately independent from Browse
        # selection state.  The store is created only after ROI load.
        self.proofread_store = None
        self.proofreading_enabled = False
        self._proof_key_filter: _ProofreadingKeyFilter | None = None
        self._proof_canvas_native = None
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target: tuple[int, int] | None = None
        self._proof_target_zyx: tuple[float, float, float] | None = None
        self._proof_target_volume_index: int | None = None
        self._proof_target_context: tuple[int, int, int] | None = None
        self._proof_mouse_callback_installed = False

        self._setup_ui()
        # Initial spin-box values are defaults, not an unapplied user draft.
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target = None
        self.proofreading_toggle.setEnabled(False)
        self._connect_viewer_events()
        self._bind_keys()
        self._refresh_image_layers()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _setup_ui(self) -> None:
        outer_layout = QVBoxLayout()
        outer_layout.setContentsMargins(0, 0, 0, 0)
        self.setMinimumWidth(320)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scroll_content = QWidget()
        layout = QVBoxLayout(self.scroll_content)

        header = QLabel("Worm Neuron Annotator")
        header.setFont(QFont("Arial", 12, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        self.image_layer_group = self._build_image_layer_group()
        layout.addWidget(self.image_layer_group)
        self.orientation_group = self._build_orientation_group()
        layout.addWidget(self.orientation_group)
        layout.addWidget(self._build_z_layer_group())
        layout.addWidget(self._build_roi_group())
        layout.addWidget(self._build_proofreading_group())
        layout.addWidget(self._build_selection_group())
        layout.addWidget(self._build_annotation_group())
        layout.addWidget(self._build_status_group())
        layout.addStretch(1)
        self.scroll_area.setWidget(self.scroll_content)
        outer_layout.addWidget(self.scroll_area)
        self.setLayout(outer_layout)

    def _build_image_layer_group(self) -> QGroupBox:
        group = QGroupBox("Image Layer")
        group_layout = QVBoxLayout()
        self.image_combo = QComboBox()
        self.image_combo.setSizeAdjustPolicy(
            QComboBox.AdjustToMinimumContentsLengthWithIcon
        )
        self.image_combo.setMinimumContentsLength(10)
        self.image_combo.currentIndexChanged.connect(
            self._on_image_changed
        )
        group_layout.addWidget(QLabel("Spatial source:"))
        group_layout.addWidget(self.image_combo)
        group.setLayout(group_layout)
        # Compatibility for callers that previously used the Z-only selector.
        self.z_image_combo = self.image_combo
        return group

    def _build_orientation_group(self) -> QGroupBox:
        group = QGroupBox("Worm Orientation")
        group_layout = QGridLayout()

        group_layout.addWidget(QLabel("Rotation (clockwise):"), 0, 0)
        self.orientation_rotation_combo = QComboBox()
        for rotation in ALLOWED_ROTATIONS:
            self.orientation_rotation_combo.addItem(f"{rotation}°", rotation)
        self.orientation_rotation_combo.currentIndexChanged.connect(
            self._on_orientation_controls_changed
        )
        group_layout.addWidget(self.orientation_rotation_combo, 0, 1)

        self.flip_horizontal_checkbox = QCheckBox("Flip screen horizontal")
        self.flip_horizontal_checkbox.toggled.connect(
            self._on_orientation_controls_changed
        )
        group_layout.addWidget(self.flip_horizontal_checkbox, 1, 0, 1, 2)

        self.flip_vertical_checkbox = QCheckBox("Flip screen vertical")
        self.flip_vertical_checkbox.toggled.connect(
            self._on_orientation_controls_changed
        )
        group_layout.addWidget(self.flip_vertical_checkbox, 2, 0, 1, 2)

        self.orientation_reset_btn = QPushButton("Reset")
        self.orientation_reset_btn.clicked.connect(self.reset_orientation)
        group_layout.addWidget(self.orientation_reset_btn, 3, 0, 1, 2)
        group_layout.setColumnStretch(1, 1)

        group.setLayout(group_layout)
        group.setEnabled(False)
        return group

    def _build_z_layer_group(self) -> QGroupBox:
        group = QGroupBox("Z Layers")
        group_layout = QVBoxLayout()

        cuts_layout = QHBoxLayout()
        cuts_layout.addWidget(QLabel("Cuts:"))
        self.z_cuts_input = QLineEdit()
        self.z_cuts_input.setPlaceholderText("4,10")
        self.z_cuts_input.setToolTip(
            "Z cuts define half-open ranges [start, stop); "
            "a boundary slice belongs to the following layer."
        )
        self.z_cuts_input.textChanged.connect(self._sync_z_profile_cuts)
        cuts_layout.addWidget(self.z_cuts_input, 1)
        self.split_z_btn = QPushButton("Split")
        self.split_z_btn.clicked.connect(self.split_z_layers)
        cuts_layout.addWidget(self.split_z_btn)
        group_layout.addLayout(cuts_layout)

        profile_header = QHBoxLayout()
        self.z_profile_label = QLabel("Pixels >")
        profile_header.addWidget(self.z_profile_label)
        self.z_profile_threshold_spin = QDoubleSpinBox()
        self.z_profile_threshold_spin.setRange(-1_000_000_000, 1_000_000_000)
        self.z_profile_threshold_spin.setDecimals(3)
        self.z_profile_threshold_spin.setValue(170)
        self.z_profile_threshold_spin.setKeyboardTracking(False)
        self.z_profile_threshold_spin.setToolTip(
            "Count pixels with intensity strictly greater than this threshold."
        )
        self.z_profile_threshold_spin.valueChanged.connect(
            self.refresh_z_profile
        )
        profile_header.addWidget(self.z_profile_threshold_spin)
        self.z_profile_state_label = QLabel("")
        profile_header.addWidget(self.z_profile_state_label)
        profile_header.addStretch(1)
        self.z_profile_refresh_btn = QPushButton("Refresh")
        self.z_profile_refresh_btn.clicked.connect(self.refresh_z_profile)
        profile_header.addWidget(self.z_profile_refresh_btn)
        group_layout.addLayout(profile_header)
        self.z_profile = ZThresholdProfileWidget()
        self.z_profile.cutToggled.connect(self._toggle_z_profile_cut)
        group_layout.addWidget(self.z_profile)

        view_layout = QHBoxLayout()
        view_layout.addWidget(QLabel("Show:"))
        self.z_view_combo = QComboBox()
        self.z_view_combo.addItem("All", None)
        self.z_view_combo.setEnabled(False)
        self.z_view_combo.currentIndexChanged.connect(self._on_z_view_changed)
        view_layout.addWidget(self.z_view_combo, 1)
        self.clear_z_btn = QPushButton("Clear")
        self.clear_z_btn.setEnabled(False)
        self.clear_z_btn.clicked.connect(self.clear_z_layers)
        view_layout.addWidget(self.clear_z_btn)
        group_layout.addLayout(view_layout)

        description = QLabel("Split Image by z; synchronize neuron boxes.")
        description.setToolTip(
            "Neuron boxes are assigned to one layer by center Z."
        )
        group_layout.addWidget(description)
        group.setLayout(group_layout)
        return group

    def _build_roi_group(self) -> QGroupBox:
        group = QGroupBox("Neuron ROI Source")
        group_layout = QVBoxLayout()

        path_layout = QHBoxLayout()
        self.roi_path_input = QLineEdit()
        self.roi_path_input.setReadOnly(True)
        self.roi_path_input.setPlaceholderText("Load neuron_pt_tuple.npy")
        self.load_roi_btn = QPushButton("Load NPY")
        self.load_roi_btn.setEnabled(False)
        self.load_roi_btn.clicked.connect(self.load_roi_npy)
        self.unload_roi_btn = QPushButton("Unload")
        self.unload_roi_btn.clicked.connect(self.unload_roi)
        self.unload_roi_btn.setEnabled(False)
        path_layout.addWidget(self.roi_path_input, 1)
        path_layout.addWidget(self.load_roi_btn)
        path_layout.addWidget(self.unload_roi_btn)
        group_layout.addLayout(path_layout)

        config_layout = QGridLayout()
        config_layout.addWidget(QLabel("Z divisor:"), 0, 0)
        self.z_divisor_spin = QDoubleSpinBox()
        self.z_divisor_spin.setDecimals(3)
        self.z_divisor_spin.setRange(0.001, 1_000_000)
        self.z_divisor_spin.setValue(5.0)
        config_layout.addWidget(self.z_divisor_spin, 0, 1)

        config_layout.addWidget(QLabel("Volume start:"), 1, 0)
        self.volume_start_spin = QSpinBox()
        self.volume_start_spin.setRange(0, 1_000_000)
        config_layout.addWidget(self.volume_start_spin, 1, 1)

        config_layout.addWidget(QLabel("Stride:"), 2, 0)
        self.volume_stride_spin = QSpinBox()
        self.volume_stride_spin.setRange(1, 1_000_000)
        self.volume_stride_spin.setValue(1)
        config_layout.addWidget(self.volume_stride_spin, 2, 1)
        config_layout.setColumnStretch(1, 1)
        group_layout.addLayout(config_layout)

        self.roi_info_label = QLabel("No ROI loaded")
        self.roi_info_label.setWordWrap(True)
        group_layout.addWidget(self.roi_info_label)
        group.setLayout(group_layout)
        return group

    def _build_proofreading_group(self) -> QGroupBox:
        """Build the opt-in 2D proofreading controls.

        The controls are present even before an ROI is loaded so the state is
        discoverable, but remain disabled until a compatible 2D/All view and a
        ProofreadStore are available.
        """
        group = QGroupBox("Proofreading")
        layout = QVBoxLayout()

        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Proofreading:"))
        self.proofreading_toggle = QCheckBox("Off / On")
        self.proofreading_toggle.setChecked(False)
        self.proofreading_toggle.setToolTip(
            "Select the source Image layer, then enable 2D proofreading."
        )
        self.proofreading_toggle.toggled.connect(self._set_proofreading_enabled)
        mode_row.addWidget(self.proofreading_toggle)
        mode_row.addStretch(1)
        layout.addLayout(mode_row)

        sizes = QGridLayout()
        self.proof_width_spin = QDoubleSpinBox()
        self.proof_height_spin = QDoubleSpinBox()
        self.proof_depth_spin = QDoubleSpinBox()
        for spin in (
            self.proof_width_spin,
            self.proof_height_spin,
            self.proof_depth_spin,
        ):
            spin.setRange(0.001, 1_000_000)
            spin.setDecimals(3)
            spin.setSingleStep(1.0)
            spin.setKeyboardTracking(False)
            spin.valueChanged.connect(self._on_proof_size_draft_changed)
        self.proof_width_spin.setValue(7)
        self.proof_height_spin.setValue(7)
        self.proof_depth_spin.setValue(3)
        sizes.addWidget(QLabel("width:"), 0, 0)
        sizes.addWidget(self.proof_width_spin, 0, 1)
        sizes.addWidget(QLabel("height:"), 0, 2)
        sizes.addWidget(self.proof_height_spin, 0, 3)
        sizes.addWidget(QLabel("depth:"), 0, 4)
        sizes.addWidget(self.proof_depth_spin, 0, 5)
        sizes.setColumnStretch(1, 1)
        sizes.setColumnStretch(3, 1)
        sizes.setColumnStretch(5, 1)
        layout.addLayout(sizes)

        self.proof_current_box_label = QLabel("No active neuron")
        self.proof_current_box_label.setWordWrap(True)
        self.proof_current_box_label.setObjectName("proofCurrentBoxLabel")
        layout.addWidget(self.proof_current_box_label)

        self.proof_apply_size_btn = QPushButton("Apply size")
        self.proof_apply_size_btn.clicked.connect(self._proof_apply_size)
        layout.addWidget(self.proof_apply_size_btn)

        io_row = QHBoxLayout()
        self.proof_save_btn = QPushButton("Save proof edits")
        self.proof_load_btn = QPushButton("Load proof edits")
        self.proof_discard_btn = QPushButton("Discard edits")
        self.proof_save_btn.clicked.connect(self.save_proof_edits)
        self.proof_load_btn.clicked.connect(self.load_proof_edits)
        self.proof_discard_btn.clicked.connect(self.discard_proof_edits)
        io_row.addWidget(self.proof_save_btn)
        io_row.addWidget(self.proof_load_btn)
        io_row.addWidget(self.proof_discard_btn)
        layout.addLayout(io_row)

        export_row = QHBoxLayout()
        self.proof_export_btn = QPushButton("Export corrected NPY")
        self.proof_delete_all_btn = QPushButton("Delete active neuron (all volumes)")
        self.proof_retire_btn = QPushButton("Retire added neuron")
        self.proof_export_btn.clicked.connect(self.export_corrected_npy)
        self.proof_delete_all_btn.clicked.connect(self._proof_delete_all)
        self.proof_retire_btn.clicked.connect(self._proof_retire_added_neuron)
        export_row.addWidget(self.proof_export_btn)
        export_row.addWidget(self.proof_delete_all_btn)
        layout.addLayout(export_row)
        layout.addWidget(self.proof_retire_btn)

        self.proof_help_label = QLabel(
            "Click to lock target · F7 delete · F8 place · F9 add · F12 exit"
        )
        self.proof_help_label.setToolTip(
            "Select the source Image layer. In 2D/All, click to lock the cyan crosshair, then use F8 or F9."
        )
        layout.addWidget(self.proof_help_label)
        group.setLayout(layout)
        self._set_proofreading_controls_enabled(False)
        # Initial spin-box values are defaults, not an unapplied user draft.
        self._proof_size_draft_dirty = False
        return group

    # ------------------------------------------------------------------
    # Proofreading mode, cursor operations, and persistence adapters
    # ------------------------------------------------------------------
    def _proof_view_allowed(self) -> bool:
        """Return whether the current viewer state supports proofreading."""
        return bool(
            self.roi_dataset is not None
            and self.current_image is not None
            and self.current_image in self.viewer.layers
            and not self._proof_detached
            and self.viewer.dims.ndisplay == 2
            and self._view_axes_supported()
            and self._active_z_range() is None
            and self._current_volume_index() is not None
            and self._image_matches_proof_session(self.current_image)
            and ProofreadStore is not None
        )

    def _proof_interaction_allowed(self) -> bool:
        """Return whether a proof edit may consume the current canvas state."""
        return bool(
            self._proof_view_allowed()
            and self.viewer.layers.selection.active is self.current_image
        )

    def _set_proofreading_controls_enabled(self, enabled: bool) -> None:
        for control in (
            self.proof_apply_size_btn,
            self.proof_delete_all_btn,
            self.proof_retire_btn,
            self.proof_width_spin,
            self.proof_height_spin,
            self.proof_depth_spin,
        ):
            control.setEnabled(bool(enabled))
        session_available = self.proofread_store is not None
        for control in (
            self.proof_save_btn,
            self.proof_load_btn,
            self.proof_discard_btn,
            self.proof_export_btn,
        ):
            control.setEnabled(session_available)

    def _set_proofreading_enabled(self, enabled: bool) -> None:
        requested = bool(enabled)
        if requested and self.viewer.layers.selection.active is not self.current_image:
            self.proofreading_toggle.blockSignals(True)
            self.proofreading_toggle.setChecked(False)
            self.proofreading_toggle.blockSignals(False)
            self.proofreading_enabled = False
            self._remove_proof_key_filter()
            self._remove_proof_mouse_callback()
            self._clear_proof_target()
            self._set_proofreading_controls_enabled(False)
            self.update_status("Select the source Image layer first", "orange")
            return
        if requested and not self._proof_view_allowed():
            self.proofreading_toggle.blockSignals(True)
            self.proofreading_toggle.setChecked(False)
            self.proofreading_toggle.blockSignals(False)
            self.proofreading_enabled = False
            self._remove_proof_key_filter()
            self._remove_proof_mouse_callback()
            self._clear_proof_target()
            self._set_proofreading_controls_enabled(False)
            self.update_status(
                "Proofreading requires a loaded ROI, 2D view, and Z view All",
                "orange",
            )
            return
        self.proofreading_enabled = requested
        if requested:
            if not self._install_proof_key_filter():
                self.proofreading_enabled = False
                self.proofreading_toggle.blockSignals(True)
                self.proofreading_toggle.setChecked(False)
                self.proofreading_toggle.blockSignals(False)
                self._set_proofreading_controls_enabled(False)
                return
            self._install_proof_mouse_callback()
            self._ensure_proof_target_layer()
        else:
            self._remove_proof_key_filter()
            self._remove_proof_mouse_callback()
            self._clear_proof_target()
        self._set_proofreading_controls_enabled(
            self.proofreading_enabled and self._proof_view_allowed()
        )
        self._update_proof_size_controls()
        self._update_info()

    def _install_proof_key_filter(self) -> bool:
        if self._proof_key_filter is not None:
            return True
        canvas = self._get_proof_canvas_native()
        if canvas is None:
            self.update_status("Canvas hotkeys are unavailable", "orange")
            return False
        if not hasattr(canvas, "installEventFilter"):
            self.update_status("Canvas hotkeys are unavailable", "orange")
            return False
        self._proof_canvas_native = canvas
        self._proof_key_filter = _ProofreadingKeyFilter(self)
        canvas.installEventFilter(self._proof_key_filter)
        return True

    def _get_proof_canvas_native(self):
        """Return the canvas widget without using napari's deprecated accessor.

        napari 0.8 exposes the Qt viewer through ``Window._qt_viewer``.  It is
        technically private, but is the only available canvas-native bridge in
        the supported release range (``napari>=0.8,<0.9``).  Keep this access
        isolated so a future napari migration has one small compatibility
        point.  Do not fall back to ``Window.qt_viewer``: merely reading that
        public property emits the deprecation warning on napari 0.8.
        """
        try:
            window = self.viewer.window
            qt_viewer = getattr(window, "_qt_viewer", None)
            canvas = getattr(qt_viewer, "canvas", None)
            return getattr(canvas, "native", None)
        except (AttributeError, RuntimeError, TypeError):
            return None

    def _remove_proof_key_filter(self) -> None:
        filt = self._proof_key_filter
        canvas = self._proof_canvas_native
        if filt is not None and canvas is not None:
            with suppress(RuntimeError, AttributeError):
                canvas.removeEventFilter(filt)
        if filt is not None:
            filt.deleteLater()
        self._proof_key_filter = None
        self._proof_canvas_native = None

    def _install_proof_mouse_callback(self) -> None:
        """Install the viewer-level click callback exactly once."""
        callback = self._on_proof_mouse_drag
        if callback not in self.viewer.mouse_drag_callbacks:
            self.viewer.mouse_drag_callbacks.append(callback)
        self._proof_mouse_callback_installed = True

    def _remove_proof_mouse_callback(self) -> None:
        """Remove the proofreading click callback without touching layers."""
        callback = self._on_proof_mouse_drag
        while callback in self.viewer.mouse_drag_callbacks:
            self.viewer.mouse_drag_callbacks.remove(callback)
        self._proof_mouse_callback_installed = False

    def _on_proof_mouse_drag(self, viewer, event):
        """Lock one target from an unmodified left-button short click."""
        del viewer
        if (
            not self.proofreading_enabled
            or not self._proof_view_allowed()
            or self.viewer.layers.selection.active is not self.current_image
            or self.viewer.grid.enabled
            or event.type != "mouse_press"
            or event.button != 1
            or event.modifiers
        ):
            return
        try:
            press_pos = np.asarray(event.pos, dtype=float)
        except (TypeError, ValueError):
            return
        if press_pos.size != 2 or not np.all(np.isfinite(press_pos)):
            return

        valid_click = True
        yield
        while event.type == "mouse_move":
            try:
                move_pos = np.asarray(event.pos, dtype=float)
            except (TypeError, ValueError):
                move_pos = np.empty(0, dtype=float)
            if (
                event.modifiers
                or move_pos.size != 2
                or not np.all(np.isfinite(move_pos))
                or np.linalg.norm(move_pos - press_pos) > 3.0
            ):
                valid_click = False
            yield
        if (
            event.type != "mouse_release"
            or event.button != 1
            or not valid_click
            or event.modifiers
        ):
            return
        try:
            release_pos = np.asarray(event.pos, dtype=float)
        except (TypeError, ValueError):
            return
        if (
            release_pos.size != 2
            or not np.all(np.isfinite(release_pos))
            or np.linalg.norm(release_pos - press_pos) > 3.0
        ):
            return
        self._lock_proof_target(event.position)

    def _proof_world_to_zyx(
        self, world: tuple[float, ...] | np.ndarray
    ) -> tuple[float, float, float] | None:
        """Convert a canvas world coordinate to the current Image z/y/x."""
        if self.current_image is None:
            return None
        try:
            data = np.asarray(self.current_image.world_to_data(world), dtype=float)
        except (AttributeError, TypeError, ValueError, RuntimeError):
            return None
        if data.size < 3:
            return None
        zyx = np.asarray(data[-3:], dtype=float)
        zyx[0] = float(self._viewer_z())
        shape = np.asarray(self._shape_zyx(), dtype=float)
        if not np.all(np.isfinite(zyx)) or np.any(zyx < 0) or np.any(zyx >= shape):
            return None
        return tuple(float(value) for value in zyx)

    def _lock_proof_target(self, world) -> None:
        """Lock and render a proofreading target for this volume and Z slice."""
        if not self.proofreading_enabled or not self._proof_interaction_allowed():
            return
        center = self._proof_world_to_zyx(world)
        volume_index = self._current_volume_index()
        if center is None or volume_index is None:
            return
        viewer_t = self._viewer_time()
        viewer_z = self._viewer_z()
        if self._ensure_proof_target_layer() is None:
            return
        self._proof_target_zyx = center
        self._proof_target_volume_index = volume_index
        self._proof_target_context = (volume_index, viewer_t, viewer_z)
        self._set_proof_target_layer_data(center)
        self.update_status("Proofreading target locked", "green")

    def _clear_proof_target(self) -> None:
        """Clear the session-only target and its crosshair geometry."""
        self._proof_target_zyx = None
        self._proof_target_volume_index = None
        self._proof_target_context = None
        layer = self._managed_proof_target_layer()
        if layer is not None:
            layer.data = np.empty((0, 2, layer.ndim), dtype=float)
            # Assigning empty Vectors data resets napari's editable flag.
            layer.editable = False
            layer.visible = False

    def _set_proof_target_layer_data(
        self, center_zyx: tuple[float, float, float]
    ) -> None:
        layer = self._managed_proof_target_layer()
        if layer is None or self.current_image is None:
            return
        # Vectors use ``[start, delta]`` records.  Two orthogonal segments
        # make a true crosshair while leaving visual length and edge width
        # independently controllable (unlike a Points ``symbol='cross'``).
        if self.current_image.ndim == 4:
            center = np.asarray(
                (float(self._viewer_time()), *center_zyx), dtype=float
            )
            y_axis, x_axis = 2, 3
        else:
            center = np.asarray(center_zyx, dtype=float)
            y_axis, x_axis = 1, 2
        half = float(PROOF_TARGET_HALF_LENGTH)
        segments = np.zeros((2, 2, self.current_image.ndim), dtype=float)
        horizontal_start = center.copy()
        horizontal_start[x_axis] -= half
        segments[0, 0] = horizontal_start
        segments[0, 1, x_axis] = 2.0 * half
        vertical_start = center.copy()
        vertical_start[y_axis] -= half
        segments[1, 0] = vertical_start
        segments[1, 1, y_axis] = 2.0 * half
        layer.data = segments
        layer.editable = False
        layer.visible = True

    def _proof_locked_target(self) -> tuple[float, float, float] | None:
        """Return the target only while it belongs to the current volume/Z."""
        volume_index = self._current_volume_index()
        context = self._proof_target_context
        if (
            self._proof_target_zyx is None
            or volume_index is None
            or context != (volume_index, self._viewer_time(), self._viewer_z())
        ):
            self._clear_proof_target()
            return None
        return self._proof_target_zyx

    def _proof_key_focus_blocked(self) -> bool:
        focus = QApplication.focusWidget()
        if focus is None:
            return False
        blocked_types = (
            QLineEdit,
            QTextEdit,
            QTreeWidget,
            QTableWidget,
            QAbstractSpinBox,
        )
        if isinstance(focus, blocked_types):
            return True
        # Qt item editors are transient children of the table/tree.
        parent = focus.parentWidget()
        while parent is not None:
            if isinstance(parent, blocked_types):
                return True
            parent = parent.parentWidget()
        return False

    def _current_volume_index(self) -> int | None:
        """Map the current Image time to the raw NPY first-axis index."""
        if self.roi_dataset is None or self.current_image is None:
            return None
        viewer_t = self._viewer_time()
        index = (
            self.volume_start_spin.value()
            + viewer_t * self.volume_stride_spin.value()
        )
        raw_t = getattr(self.roi_dataset, "time_count", None)
        if raw_t is None:
            raw_t = getattr(self.roi_dataset, "raw_T", 0)
        return int(index) if 0 <= int(index) < int(raw_t) else None

    def _proof_resolve(self, volume_index: int, neuron_id: int):
        store = self.proofread_store
        if store is None:
            return None
        return store.resolve(volume_index, neuron_id)

    def _proof_draft_size(self) -> tuple[float, float, float]:
        """Return the visible width/height/depth draft in z/y/x order."""
        return (
            self.proof_depth_spin.value(),
            self.proof_height_spin.value(),
            self.proof_width_spin.value(),
        )

    def _proof_placement_size(
        self, neuron_id: int, volume_index: int | None = None
    ) -> tuple[float, float, float]:
        store = self.proofread_store
        if store is None:
            return (3.0, 7.0, 7.0)
        if volume_index is None:
            volume_index = self._current_volume_index()
        return tuple(
            float(value)
            for value in store.size_for_placement(neuron_id, volume_index)
        )

    def _update_proof_current_box_status(self) -> None:
        """Show the active box center, Image time, and edit marker.

        The proofreading store addresses observations by raw ``volume_index``;
        the label intentionally reports the user-facing Image ``t`` so it
        matches the napari view.  ``NeuronBox.center_zyx`` is reordered to
        user-facing ``x, y, z`` here; the internal geometry remains z/y/x.
        """
        label = getattr(self, "proof_current_box_label", None)
        if label is None:
            return

        active_id = self.active_id
        volume_index = self._current_volume_index()
        viewer_t = self._viewer_time()
        if active_id is None:
            label.setText("No active neuron")
            return
        if volume_index is None:
            label.setText(f"Neuron {active_id}: no box, t={viewer_t}")
            return

        store = self.proofread_store
        try:
            if store is not None:
                box = store.resolve(volume_index, active_id)
                modified = (volume_index, active_id) in store.modified_observations
            elif self.roi_dataset is not None:
                box = self.roi_dataset.get_box(self._viewer_time(), active_id)
                modified = False
            else:
                box = None
                modified = False
        except (AttributeError, IndexError, RuntimeError, TypeError, ValueError):
            # A transient Image/ROI transition can invalidate the target
            # between the volume lookup and store resolution.  Keep the UI
            # informative instead of allowing a Qt callback to fail.
            box = None
            modified = False

        if box is None:
            marker = " (modified)" if modified else ""
            label.setText(f"Neuron {active_id}{marker}: no box, t={viewer_t}")
            return

        z, y, x = (float(value) for value in box.center_zyx)
        marker = " (modified)" if modified else ""
        text = (
            f"Neuron {active_id}{marker}: "
            f"center (x={x:.2f}, y={y:.2f}, z={z:.2f}), t={viewer_t}"
        )
        label.setText(text)

    def _on_proof_size_draft_changed(self, value: float) -> None:
        del value
        volume_index = self._current_volume_index()
        if self.active_id is None or volume_index is None:
            return
        self._proof_size_draft_dirty = True
        self._proof_size_draft_target = (volume_index, self.active_id)
        if hasattr(self, "info_text"):
            self._update_info()

    def _update_proof_size_controls(self) -> None:
        self._update_proof_current_box_status()
        # A target transition must explicitly Apply/Discard/Cancel this draft.
        # Never let a routine refresh silently overwrite it.
        if self._proof_size_draft_dirty:
            return
        volume_index = self._current_volume_index()
        if (
            self.active_id is None
            or self.proofread_store is None
            or volume_index is None
        ):
            size = (3.0, 7.0, 7.0)
        else:
            box = self._proof_resolve(volume_index, self.active_id)
            size = (
                box.size_zyx
                if box is not None
                else self._proof_placement_size(self.active_id, volume_index)
            )
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target = None
        self.proof_depth_spin.blockSignals(True)
        self.proof_height_spin.blockSignals(True)
        self.proof_width_spin.blockSignals(True)
        try:
            self.proof_depth_spin.setValue(max(0.001, float(size[0])))
            self.proof_height_spin.setValue(max(0.001, float(size[1])))
            self.proof_width_spin.setValue(max(0.001, float(size[2])))
        finally:
            self.proof_depth_spin.blockSignals(False)
            self.proof_height_spin.blockSignals(False)
            self.proof_width_spin.blockSignals(False)

    def _proof_apply_size(self) -> None:
        draft_target = self._proof_size_draft_target
        target_id = (
            draft_target[1]
            if self._proof_size_draft_dirty and draft_target is not None
            else self.active_id
        )
        if (
            not self.proofreading_enabled
            or not self._proof_interaction_allowed()
            or target_id is None
        ):
            return
        store = self.proofread_store
        if store is None:
            return
        size = self._proof_draft_size()
        if any(not np.isfinite(v) or v <= 0 for v in size):
            self.update_status("Box dimensions must be positive", "orange")
            return
        try:
            store.apply_size(target_id, size)
        except (AttributeError, TypeError, ValueError, RuntimeError) as error:
            self.update_status(f"Could not apply size: {error}", "red")
            return
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target = None
        self._refresh_available_ids(select_first=False)
        self._refresh_roi_layers()
        self._update_proof_current_box_status()
        self.update_status(f"Applied size to neuron {target_id}", "green")

    def _confirm_proof_size_draft(self, action: str) -> bool:
        """Resolve an unapplied size before changing its neuron/volume target."""
        if not self._proof_size_draft_dirty:
            return True
        target = self._proof_size_draft_target
        dialog = QMessageBox(self)
        dialog.setIcon(QMessageBox.Warning)
        dialog.setWindowTitle("Unapplied box size")
        dialog.setText(f"Apply the current size draft before {action}?")
        apply_button = dialog.addButton("Apply", QMessageBox.AcceptRole)
        discard_button = dialog.addButton(
            "Discard draft", QMessageBox.DestructiveRole
        )
        cancel_button = dialog.addButton(QMessageBox.Cancel)
        dialog.setDefaultButton(cancel_button)
        dialog.exec()
        clicked = dialog.clickedButton()
        if clicked is cancel_button or clicked is None:
            return False
        if clicked is apply_button:
            if self.proofread_store is None or target is None:
                self.update_status("The size draft has no valid target", "red")
                return False
            try:
                self.proofread_store.apply_size(target[1], self._proof_draft_size())
            except (TypeError, ValueError, RuntimeError) as error:
                self.update_status(f"Could not apply size: {error}", "red")
                return False
        elif clicked is not discard_button:
            return False
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target = None
        return True

    def _proof_delete_current(self) -> None:
        if (
            not self.proofreading_enabled
            or not self._proof_interaction_allowed()
            or self.active_id is None
        ):
            return
        volume_index = self._current_volume_index()
        if volume_index is None:
            return
        box = self._proof_resolve(volume_index, self.active_id)
        if box is None:
            self.update_status("Current observation is already missing", "orange")
            return
        try:
            self.proofread_store.set_observation_deleted(volume_index, self.active_id)
        except (AttributeError, TypeError, ValueError, RuntimeError) as error:
            self.update_status(f"Could not delete observation: {error}", "red")
            return
        self._refresh_after_proof_edit()
        self.update_status("Deleted current observation", "green")

    def _proof_place_cursor(self) -> None:
        if (
            not self.proofreading_enabled
            or not self._proof_interaction_allowed()
            or self.active_id is None
        ):
            return
        volume_index = self._current_volume_index()
        center = self._proof_locked_target()
        if volume_index is None or center is None:
            self.update_status(
                "Click the Image to lock a proofreading target first",
                "orange",
            )
            return
        if self._proof_resolve(volume_index, self.active_id) is not None:
            self.update_status("Current observation exists; F8 did nothing", "orange")
            return
        size = self._proof_placement_size(self.active_id, volume_index)
        try:
            self.proofread_store.set_observation_present(
                volume_index,
                self.active_id,
                center_zyx=center,
                size_zyx=size,
            )
        except TypeError:
            # Compatibility with a positional box-oriented implementation.
            from ._roi import NeuronBox

            box = NeuronBox(self.active_id, volume_index, center, size)
            self.proofread_store.set_observation_present(volume_index, self.active_id, box)
        except (AttributeError, ValueError, RuntimeError) as error:
            self.update_status(f"Could not place observation: {error}", "red")
            return
        self._clear_proof_target()
        self._refresh_after_proof_edit()
        self.update_status("Placed center at locked target", "green")

    def _proof_add_neuron(self) -> None:
        if not self.proofreading_enabled or not self._proof_interaction_allowed():
            return
        if not self._confirm_proof_size_draft("adding a neuron"):
            return
        volume_index = self._current_volume_index()
        center = self._proof_locked_target()
        if volume_index is None or center is None or self.proofread_store is None:
            self.update_status(
                "Click the Image to lock a proofreading target first",
                "orange",
            )
            return
        size = (3.0, 7.0, 7.0)
        try:
            neuron_id = self.proofread_store.add_neuron(
                volume_index, center, size_zyx=size
            )
        except TypeError:
            try:
                neuron_id = self.proofread_store.add_neuron(volume_index, center)
            except (AttributeError, ValueError, RuntimeError) as error:
                self.update_status(f"Could not add neuron: {error}", "red")
                return
        except (AttributeError, ValueError, RuntimeError) as error:
            self.update_status(f"Could not add neuron: {error}", "red")
            return
        self._clear_proof_target()
        self._available_ids = list(self.proofread_store.neuron_ids)
        self.active_id = int(neuron_id)
        self.checked_ids.add(self.active_id)
        self._refresh_available_ids(select_first=False)
        self._selection_changed(locate=False)
        self._refresh_after_proof_edit()
        self.update_status(f"Added neuron {self.active_id}", "green")

    def _proof_cancel_or_exit(self) -> None:
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target = None
        self.proofreading_toggle.blockSignals(True)
        self.proofreading_toggle.setChecked(False)
        self.proofreading_toggle.blockSignals(False)
        self._set_proofreading_enabled(False)
        self.update_status("Proofreading off", "green")

    def _set_proofreading_off_preserve_draft(self) -> None:
        """Disable proofreading without discarding an existing size draft."""
        draft_dirty = self._proof_size_draft_dirty
        draft_target = self._proof_size_draft_target
        self.proofreading_toggle.blockSignals(True)
        self.proofreading_toggle.setChecked(False)
        self.proofreading_toggle.blockSignals(False)
        self._set_proofreading_enabled(False)
        self._proof_size_draft_dirty = draft_dirty
        self._proof_size_draft_target = draft_target

    def _proof_delete_all(self) -> None:
        if (
            not self.proofreading_enabled
            or not self._proof_interaction_allowed()
            or self.active_id is None
            or self.proofread_store is None
        ):
            return
        answer = QMessageBox.question(
            self,
            "Delete all observations",
            f"Delete neuron {self.active_id} from all volumes?\nThe ID will be retained.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        volume_index = self._current_volume_index()
        placement_size = self._proof_placement_size(
            self.active_id, volume_index
        )
        try:
            self.proofread_store.delete_all_observations(self.active_id)
            self.proofread_store.placement_size[self.active_id] = placement_size
        except (AttributeError, ValueError, RuntimeError) as error:
            self.update_status(f"Could not delete all observations: {error}", "red")
            return
        self._refresh_after_proof_edit()
        self.update_status(f"Deleted neuron {self.active_id} in all volumes", "green")

    def _proof_retire_added_neuron(self) -> None:
        store = self.proofread_store
        neuron_id = self.active_id
        if (
            not self.proofreading_enabled
            or not self._proof_interaction_allowed()
            or store is None
            or neuron_id is None
            or neuron_id < store.raw_N
            or neuron_id in store.retired_ids
        ):
            return
        if not self._confirm_proof_size_draft("retiring an added neuron"):
            return
        answer = QMessageBox.question(
            self,
            "Retire added neuron",
            f"Retire added neuron {neuron_id}?\n"
            "Its numeric ID will remain reserved and will not be reused.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        try:
            store.retire_added_neuron(neuron_id)
        except (AttributeError, TypeError, ValueError, RuntimeError) as error:
            self.update_status(f"Could not retire neuron: {error}", "red")
            return
        self._proof_size_draft_dirty = False
        self._proof_size_draft_target = None
        self._refresh_available_ids(select_first=False)
        self._refresh_after_proof_edit()
        self.update_status(f"Retired added neuron {neuron_id}", "green")

    def _refresh_after_proof_edit(self) -> None:
        self._refresh_selection_item_styles()
        self._refresh_roi_layers()
        self._update_proof_size_controls()
        self._update_info()

    def save_proof_edits(self) -> None:
        store = self.proofread_store
        if store is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save proof edits", self._proof_default_path(),
            "Proof sidecar (*.json);;All files (*)",
        )
        if not path:
            return
        try:
            store.save(path)
        except (OSError, PermissionError, ValueError, RuntimeError) as error:
            self.update_status(f"Proof save failed: {error}", "red")
            return
        self._refresh_available_ids(select_first=False)
        self._refresh_after_proof_edit()
        self.update_status(f"Saved proof edits: {Path(path).name}", "green")

    def _save_proof_edits_for_transition(self) -> bool:
        """Save pending proof edits and report whether a transition may run."""
        store = self.proofread_store
        if store is None or not store.dirty:
            return True
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save proof edits",
            self._proof_default_path(),
            "Proof sidecar (*.json);;All files (*)",
        )
        if not path:
            return False
        try:
            store.save(path)
        except (OSError, PermissionError, ValueError, RuntimeError) as error:
            self.update_status(f"Proof save failed: {error}", "red")
            return False
        self._refresh_available_ids(select_first=False)
        self._refresh_after_proof_edit()
        self.update_status(f"Saved proof edits: {Path(path).name}", "green")
        return True

    def _confirm_proof_transition(self, action: str) -> bool:
        """Resolve draft/dirty state before discarding the current session."""
        store = self.proofread_store
        if not self._confirm_proof_size_draft(action):
            return False

        if store is None or not store.dirty:
            return True
        result = QMessageBox.warning(
            self,
            "Unsaved proofreading edits",
            f"Save proofreading edits before {action}?",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            QMessageBox.Cancel,
        )
        if result == QMessageBox.Cancel:
            return False
        if result == QMessageBox.Save:
            return self._save_proof_edits_for_transition()
        if result == QMessageBox.Discard:
            store.discard()
            self._refresh_available_ids(select_first=False)
            return True
        return False

    def load_proof_edits(self) -> None:
        store_type = ProofreadStore
        if store_type is None or self.roi_dataset is None:
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Load proof edits", self._proof_default_path(),
            "Proof sidecar (*.json);;All files (*)",
        )
        if not path:
            return
        try:
            loaded = store_type(
                self.roi_dataset,
                image_signature=self._proof_image_signature(
                    self.current_image
                ),
            )
            loaded.load(path)
        except (OSError, PermissionError, ValueError, RuntimeError, TypeError) as error:
            self.update_status(f"Proof load failed: {error}", "red")
            return
        if not self._confirm_proof_transition("loading another proof sidecar"):
            return
        self.proofread_store = loaded
        self._refresh_available_ids(select_first=False)
        self._refresh_after_proof_edit()
        self.update_status(f"Loaded proof edits: {Path(path).name}", "green")

    def discard_proof_edits(self) -> None:
        store = self.proofread_store
        if store is None:
            return
        try:
            store.discard()
        except (AttributeError, RuntimeError, ValueError) as error:
            self.update_status(f"Could not discard proof edits: {error}", "red")
            return
        self._refresh_available_ids(select_first=False)
        self._refresh_after_proof_edit()
        self.update_status("Discarded unsaved proof edits", "green")

    def export_corrected_npy(self) -> None:
        store = self.proofread_store
        if store is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export corrected NPY", "", "NumPy arrays (*.npy);;All files (*)"
        )
        if not path:
            return
        try:
            store.export_corrected_npy(path)
        except (OSError, PermissionError, ValueError, RuntimeError) as error:
            self.update_status(f"NPY export failed: {error}", "red")
            return
        self.update_status(f"Exported corrected NPY: {Path(path).name}", "green")

    def _proof_default_path(self) -> str:
        if self.roi_dataset is not None and getattr(self.roi_dataset, "path", None):
            return str(Path(self.roi_dataset.path).with_suffix(".proofread.json"))
        return ""

    def _build_selection_group(self) -> QGroupBox:
        group = QGroupBox("Neuron Selection")
        group_layout = QVBoxLayout()

        self.navigation_help_label = QLabel("Q/W: last/next")
        self.navigation_help_label.setToolTip(
            "Shift+Q/W: previous/next checked neuron"
        )
        group_layout.addWidget(self.navigation_help_label)

        self.neuron_search_input = QLineEdit()
        self.neuron_search_input.setPlaceholderText(
            "Search digital ID or biological name"
        )
        self.neuron_search_input.setClearButtonEnabled(True)
        self.neuron_search_input.setToolTip(
            "Use commas for multiple IDs or biological-name fragments."
        )
        self.neuron_search_input.textChanged.connect(
            self._on_search_text_changed
        )
        self.neuron_search_input.returnPressed.connect(
            self._activate_next_search_match
        )
        group_layout.addWidget(self.neuron_search_input)

        search_results_layout = QHBoxLayout()
        self.search_matches_label = QLabel()
        self.check_matches_btn = QPushButton("Check matches")
        self.check_matches_btn.setEnabled(False)
        self.check_matches_btn.clicked.connect(self.check_search_matches)
        search_results_layout.addWidget(self.search_matches_label)
        search_results_layout.addStretch(1)
        search_results_layout.addWidget(self.check_matches_btn)
        group_layout.addLayout(search_results_layout)

        controls_layout = QHBoxLayout()
        self.check_all_btn = QPushButton("All")
        self.check_all_btn.clicked.connect(self.check_all)
        self.check_none_btn = QPushButton("None")
        self.check_none_btn.clicked.connect(self.check_none)
        controls_layout.addWidget(self.check_all_btn)
        controls_layout.addWidget(self.check_none_btn)
        controls_layout.addStretch(1)
        group_layout.addLayout(controls_layout)

        self.show_box_labels_checkbox = QCheckBox(
            "Show selected box labels"
        )
        self.show_box_labels_checkbox.toggled.connect(
            self._refresh_roi_layers
        )
        group_layout.addWidget(self.show_box_labels_checkbox)

        box_label_mode_layout = QHBoxLayout()
        box_label_mode_layout.addWidget(QLabel("Label text:"))
        self.box_label_mode_combo = QComboBox()
        self.box_label_mode_combo.addItem(
            "Biological", LABEL_MODE_BIOLOGICAL
        )
        self.box_label_mode_combo.addItem("Digital", LABEL_MODE_DIGITAL)
        self.box_label_mode_combo.addItem(
            "Digital + biological", LABEL_MODE_DIGITAL_BIOLOGICAL
        )
        self.box_label_mode_combo.currentIndexChanged.connect(
            self._refresh_roi_layers
        )
        box_label_mode_layout.addWidget(self.box_label_mode_combo, 1)
        group_layout.addLayout(box_label_mode_layout)

        box_label_color_layout = QHBoxLayout()
        box_label_color_layout.addWidget(QLabel("Text color:"))
        self.box_label_color_btn = QPushButton()
        self.box_label_color_btn.setToolTip(
            "Choose the color of selected neuron box text."
        )
        self.box_label_color_btn.clicked.connect(
            self._choose_box_label_color
        )
        box_label_color_layout.addWidget(self.box_label_color_btn)
        box_label_color_layout.addStretch(1)
        group_layout.addLayout(box_label_color_layout)
        self._update_box_label_color_button()

        self.selection_tree = QTreeWidget()
        self.selection_tree.setColumnCount(2)
        self.selection_tree.setHeaderLabels(["", "Neuron"])
        self.selection_tree.setRootIsDecorated(False)
        self.selection_tree.setAlternatingRowColors(False)
        self.selection_tree.setSelectionMode(QAbstractItemView.SingleSelection)
        self.selection_tree.setMinimumHeight(190)
        self.selection_tree.header().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.selection_tree.header().setSectionResizeMode(
            1, QHeaderView.Stretch
        )
        self.selection_tree.itemChanged.connect(
            self._on_selection_item_changed
        )
        self.selection_tree.itemClicked.connect(
            self._on_selection_item_clicked
        )
        group_layout.addWidget(self.selection_tree)
        group.setLayout(group_layout)
        return group

    def _choose_box_label_color(self) -> None:
        color = QColorDialog.getColor(
            QColor(self._box_label_color),
            self,
            "Select box label text color",
        )
        if not color.isValid():
            return
        self._box_label_color = color.name()
        self._update_box_label_color_button()
        box_label_layer = self._managed_box_label_layer()
        if box_label_layer is not None:
            box_label_layer.text.color = self._box_label_color
            box_label_layer.refresh()

    def _update_box_label_color_button(self) -> None:
        color = QColor(self._box_label_color)
        foreground = "#000000" if color.lightness() > 127 else "#ffffff"
        self.box_label_color_btn.setText(color.name().upper())
        self.box_label_color_btn.setStyleSheet(
            f"background-color: {color.name()}; color: {foreground};"
        )

    def _build_annotation_group(self) -> QGroupBox:
        group = QGroupBox("Neuron Annotation")
        group_layout = QVBoxLayout()
        sheet_layout = QHBoxLayout()

        sheet_layout.addWidget(QLabel("Sheet:"))
        self.sheet_name_input = QLineEdit("Neuron Annotations")
        sheet_layout.addWidget(self.sheet_name_input, 1)
        group_layout.addLayout(sheet_layout)

        controls = QHBoxLayout()
        self.load_current_ids_btn = QPushButton("Current IDs")
        self.load_current_ids_btn.clicked.connect(
            self.load_current_ids_to_annotation
        )
        controls.addWidget(self.load_current_ids_btn)

        self.load_excel_btn = QPushButton("Load")
        self.load_excel_btn.clicked.connect(self.load_excel_to_annotation)
        self.load_excel_btn.setEnabled(EXCEL_AVAILABLE)
        controls.addWidget(self.load_excel_btn)
        self.save_annotation_btn = QPushButton("Save")
        self.save_annotation_btn.clicked.connect(self.save_annotation_to_excel)
        self.save_annotation_btn.setEnabled(EXCEL_AVAILABLE)
        controls.addWidget(self.save_annotation_btn)
        group_layout.addLayout(controls)

        self.annotation_table = QTableWidget(0, 3)
        self.annotation_table.setHorizontalHeaderLabels(
            ["digital", "biological", "annotation"]
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.Stretch
        )
        self.annotation_table.setSelectionBehavior(
            QAbstractItemView.SelectRows
        )
        self.annotation_table.setSelectionMode(
            QAbstractItemView.SingleSelection
        )
        self.annotation_table.itemSelectionChanged.connect(
            self._on_annotation_selection_changed
        )
        self.annotation_table.itemChanged.connect(
            self._on_annotation_item_changed
        )
        group_layout.addWidget(self.annotation_table)
        group.setLayout(group_layout)
        return group

    def _build_status_group(self) -> QGroupBox:
        group = QGroupBox("Status")
        group_layout = QVBoxLayout()
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: green;")
        group_layout.addWidget(self.status_label)
        self.info_text = QTextEdit()
        self.info_text.setReadOnly(True)
        self.info_text.setMaximumHeight(90)
        group_layout.addWidget(self.info_text)
        group.setLayout(group_layout)
        return group

    # ------------------------------------------------------------------
    # Viewer and layer lifecycle
    # ------------------------------------------------------------------
    def _connect_viewer_events(self) -> None:
        events = self.viewer.layers.events
        events.inserted.connect(self._refresh_image_layers)
        events.removed.connect(self._on_layer_removed)
        events.reordered.connect(self._refresh_image_layers)
        if hasattr(events, "renamed"):
            events.renamed.connect(self._refresh_image_layers)

        dims_events = self.viewer.dims.events
        dims_events.point.connect(self._on_dims_changed)
        dims_events.ndisplay.connect(self._on_dims_changed)
        dims_events.order.connect(self._on_dims_order_changed)
        self.viewer.camera.events.orientation.connect(
            self._on_camera_orientation_changed
        )

    def _disconnect_viewer_events(self) -> None:
        events = self.viewer.layers.events
        for emitter, callback in (
            (events.inserted, self._refresh_image_layers),
            (events.removed, self._on_layer_removed),
            (events.reordered, self._refresh_image_layers),
        ):
            with suppress(TypeError, ValueError):
                emitter.disconnect(callback)
        if hasattr(events, "renamed"):
            with suppress(TypeError, ValueError):
                events.renamed.disconnect(self._refresh_image_layers)

        dims_events = self.viewer.dims.events
        for emitter in (
            dims_events.point,
            dims_events.ndisplay,
        ):
            with suppress(TypeError, ValueError):
                emitter.disconnect(self._on_dims_changed)
        with suppress(TypeError, ValueError):
            dims_events.order.disconnect(self._on_dims_order_changed)
        with suppress(TypeError, ValueError):
            self.viewer.camera.events.orientation.disconnect(
                self._on_camera_orientation_changed
            )

    def _capture_orientation_baseline(self) -> None:
        self._orientation_baseline_order = tuple(self.viewer.dims.order)
        self._orientation_baseline_camera = self._camera_orientation2d()
        self._orientation_baseline_ndim = int(self.viewer.dims.ndim)

    def _ensure_orientation_baseline(self) -> None:
        if (
            self._orientation_baseline_order is None
            or self._orientation_baseline_camera is None
            or self._orientation_baseline_ndim != int(self.viewer.dims.ndim)
        ):
            self._rebase_orientation()

    def _rebase_orientation(self) -> None:
        self._capture_orientation_baseline()
        self._orientation_state = OrientationState()
        self._sync_orientation_controls()
        self._update_orientation_controls_enabled()

    def _camera_orientation2d(self) -> Orientation2D:
        vertical, horizontal = self.viewer.camera.orientation2d
        return (
            getattr(vertical, "value", vertical),
            getattr(horizontal, "value", horizontal),
        )

    def _on_orientation_controls_changed(self, value=None) -> None:
        del value
        if self._orientation_ui_sync or self._closed:
            return
        rotation = self.orientation_rotation_combo.currentData()
        if rotation is None:
            return
        self._orientation_state = OrientationState(
            int(rotation),
            self.flip_horizontal_checkbox.isChecked(),
            self.flip_vertical_checkbox.isChecked(),
        )
        self._apply_orientation_state()

    def _apply_orientation_state(self) -> None:
        if self.current_image is None:
            self._update_orientation_controls_enabled()
            return
        self._ensure_orientation_baseline()
        if (
            self._orientation_baseline_order is None
            or self._orientation_baseline_camera is None
        ):
            return

        ndim = int(self.viewer.dims.ndim)
        target_order, target_camera = resolve_orientation(
            self._orientation_baseline_order,
            self._orientation_baseline_camera,
            self._orientation_state,
            y_axis=ndim - 2,
            x_axis=ndim - 1,
        )
        self._write_viewer_orientation(target_order, target_camera)
        self._update_orientation_controls_enabled()

    def _write_viewer_orientation(
        self,
        order: tuple[int, ...],
        camera: Orientation2D,
    ) -> None:
        center = tuple(self.viewer.camera.center)
        zoom = float(self.viewer.camera.zoom)
        angles = tuple(self.viewer.camera.angles)
        current_step = tuple(self.viewer.dims.current_step)
        self._orientation_applying = True
        try:
            if tuple(self.viewer.dims.order) != order:
                self.viewer.dims.order = order
            if self._camera_orientation2d() != camera:
                self.viewer.camera.orientation2d = camera
            if tuple(self.viewer.dims.current_step) != current_step:
                self.viewer.dims.current_step = current_step
            if not np.allclose(self.viewer.camera.center, center):
                self.viewer.camera.center = center
            if not np.isclose(self.viewer.camera.zoom, zoom):
                self.viewer.camera.zoom = zoom
            if not np.allclose(self.viewer.camera.angles, angles):
                self.viewer.camera.angles = angles
        finally:
            self._orientation_applying = False

    def reset_orientation(self) -> None:
        """Restore the captured viewer orientation for this session."""
        if self._orientation_baseline_order is None:
            return
        self._orientation_state = OrientationState()
        self._sync_orientation_controls()
        self._apply_orientation_state()

    def _restore_orientation_baseline(self) -> None:
        if (
            self._orientation_baseline_order is None
            or self._orientation_baseline_camera is None
            or self._orientation_baseline_ndim != int(self.viewer.dims.ndim)
        ):
            return
        self._orientation_state = OrientationState()
        self._sync_orientation_controls()
        self._write_viewer_orientation(
            self._orientation_baseline_order,
            self._orientation_baseline_camera,
        )

    def _sync_orientation_controls(self) -> None:
        self._orientation_ui_sync = True
        try:
            index = self.orientation_rotation_combo.findData(
                self._orientation_state.rotation_degrees
            )
            self.orientation_rotation_combo.setCurrentIndex(index)
            self.flip_horizontal_checkbox.setChecked(
                self._orientation_state.flip_horizontal
            )
            self.flip_vertical_checkbox.setChecked(
                self._orientation_state.flip_vertical
            )
        finally:
            self._orientation_ui_sync = False

    def _update_orientation_controls_enabled(self) -> None:
        enabled = (
            self.current_image is not None
            and self.current_image in self.viewer.layers
            and self._orientation_baseline_order is not None
            and self._view_axes_supported()
        )
        self.orientation_group.setEnabled(enabled)
        self.orientation_reset_btn.setEnabled(
            enabled and not self._orientation_state.is_identity
        )

    def _on_dims_order_changed(self, event=None) -> None:
        if self._closed:
            return
        if not self._orientation_applying:
            self._rebase_orientation()
        self._on_dims_changed(event)

    def _on_camera_orientation_changed(self, event=None) -> None:
        del event
        if self._closed or self._orientation_applying:
            return
        self._rebase_orientation()

    def _bind_keys(self) -> None:
        bindings = (
            ("Q", self._previous_key),
            ("W", self._next_key),
            ("Shift-Q", self._previous_checked_key),
            ("Shift-W", self._next_checked_key),
            ("G", self._previous_z_key),
            ("H", self._next_z_key),
            ("J", self._previous_time_key),
            ("K", self._next_time_key),
            ("Shift-J", self._previous_time_fast_key),
            ("Shift-K", self._next_time_fast_key),
        )
        for key, callback in bindings:
            try:
                self.viewer.bind_key(key, callback)
                self._keys_bound.append(key)
            except ValueError:
                self.update_status(
                    f"Hotkey {key} is already assigned; existing controls remain available",
                    "orange",
                )

    def _unbind_keys(self) -> None:
        for key in self._keys_bound:
            with suppress(KeyError, ValueError):
                self.viewer.bind_key(key, None, overwrite=True)
        self._keys_bound.clear()

    def _refresh_image_layers(self, event=None) -> None:
        del event
        if self._closed:
            return
        current = self.current_image
        sources = [
            layer
            for layer in self.viewer.layers
            if isinstance(layer, Image)
            and layer.metadata.get(ROLE_KEY) != ROLE_Z_IMAGE
            and self._is_compatible_image_source(layer)
        ]

        self.image_combo.blockSignals(True)
        self.image_combo.clear()
        if sources:
            for layer in sources:
                self.image_combo.addItem(layer.name, layer)
            target = (
                current
                if current in sources
                else None
                if self._proof_detached
                else sources[0]
            )
            self.image_combo.setCurrentIndex(
                sources.index(target) if target in sources else -1
            )
            self.image_combo.setEnabled(self._z_source_image is None)
            self.split_z_btn.setEnabled(True)
        else:
            target = None
            self.image_combo.addItem("No compatible Image layers")
            self.image_combo.setEnabled(False)
            self.split_z_btn.setEnabled(False)
        self.image_combo.blockSignals(False)
        self.load_roi_btn.setEnabled(
            target is not None and self.roi_dataset is None
        )
        if target is not current:
            self._set_current_image(target)
        else:
            self._update_orientation_controls_enabled()

    def _on_image_changed(self, index: int) -> None:
        del index
        source = self.image_combo.currentData()
        previous = self.current_image
        target = source if isinstance(source, Image) else None
        if (
            target is not None
            and self._proof_detached
            and not self._image_matches_proof_session(target)
        ):
            self.update_status(
                "Selected Image does not match the detached proof session",
                "orange",
            )
            return
        if not self._set_current_image(target):
            self.image_combo.blockSignals(True)
            try:
                previous_index = self.image_combo.findData(previous)
                if previous_index >= 0:
                    self.image_combo.setCurrentIndex(previous_index)
            finally:
                self.image_combo.blockSignals(False)

    def _image_matches_proof_session(self, source: Image) -> bool:
        """Check the spatial contract before reattaching dirty proof state."""
        if source.ndim not in (3, 4):
            return False
        expected = getattr(self.proofread_store, "image_signature", None)
        if expected is None:
            return True
        return self._proof_image_signature(source) == expected

    def _proof_session_has_content(self) -> bool:
        """Return whether the current proof session carries user state."""
        store = self.proofread_store
        if store is None:
            return self._proof_size_draft_dirty
        return bool(
            self._proof_size_draft_dirty
            or store.dirty
            or store.observation_patches
            or store.delete_all_ids
            or store.placement_size
            or store.committed_added_ids
            or store.provisional_added_ids
            or store.retired_ids
        )

    @staticmethod
    def _proof_image_signature(source: Image) -> dict[str, object]:
        return {
            "shape": list(source.data.shape),
            "scale": [float(value) for value in source.scale],
            "translate": [float(value) for value in source.translate],
            "axis_labels": [str(value) for value in source.axis_labels],
            "units": [str(value) for value in source.units],
        }

    def _set_current_image(
        self, source: Image | None, *, force: bool = False
    ) -> bool:
        if source is self.current_image:
            return True
        rebind_clean_proof_store = bool(
            source is not None
            and self.roi_dataset is not None
            and self.proofread_store is not None
            and not self._proof_session_has_content()
            and not self._image_matches_proof_session(source)
        )
        if (
            source is not None
            and self.roi_dataset is not None
            and self._proof_session_has_content()
            and not self._image_matches_proof_session(source)
        ):
            self.update_status(
                "Image spatial metadata does not match the proof session",
                "orange",
            )
            return False
        if (
            not force
            and self.roi_dataset is not None
            and not self._confirm_proof_transition("switching Image")
        ):
            return False
        preserve_draft = force and self._proof_size_draft_dirty
        if self.proofreading_enabled:
            self._proof_cancel_or_exit()
        if preserve_draft:
            self._proof_size_draft_dirty = True
        self._disconnect_current_image_events()
        if self._z_ranges:
            self._clear_z_layers()
        self._remove_roi_layers()
        self.current_image = source
        self._last_viewer_time = (
            self._viewer_time()
            if source is not None and source.ndim == 4
            else None
        )
        if source is not None:
            self._proof_detached = False
            if rebind_clean_proof_store and ProofreadStore is not None:
                self.proofread_store = ProofreadStore(
                    self.roi_dataset,
                    image_signature=self._proof_image_signature(source),
                )
        self.load_roi_btn.setEnabled(
            source is not None and self.roi_dataset is None
        )
        if source is not None and not force:
            # Clearing an active Z session refreshes this selector against the
            # previous source. Re-sync it after the new authority is installed.
            self._refresh_image_layers()
        if not isinstance(source, Image):
            self._z_profile_source = None
            self._z_profile_time = None
            self.z_profile.clear_profile()
            self.z_profile_state_label.clear()
            self.z_profile_refresh_btn.setEnabled(False)
            self._refresh_selection_item_styles()
            self._update_orientation_controls_enabled()
            self._update_proof_current_box_status()
            return True
        self._validate_image_source(source)
        self._ensure_orientation_baseline()
        source.events.data.connect(self._on_current_image_changed)
        for event_name in Z_SOURCE_GEOMETRY_EVENTS:
            getattr(source.events, event_name).connect(
                self._on_current_image_changed
            )
        self.z_profile_refresh_btn.setEnabled(True)
        if source is not self._z_profile_source:
            self._z_profile_source = None
            self._z_profile_time = None
            self.z_profile.clear_profile()
            self.refresh_z_profile()
        if self.roi_dataset is not None:
            self._ensure_roi_layers()
            self._refresh_roi_layers()
        self._refresh_selection_item_styles()
        self._update_roi_info()
        self._update_proof_current_box_status()
        self._update_orientation_controls_enabled()
        return True

    def _disconnect_current_image_events(self) -> None:
        source = self.current_image
        if source is None:
            return
        with suppress(TypeError, ValueError):
            source.events.data.disconnect(self._on_current_image_changed)
        for event_name in Z_SOURCE_GEOMETRY_EVENTS:
            with suppress(TypeError, ValueError):
                getattr(source.events, event_name).disconnect(
                    self._on_current_image_changed
                )

    def _on_current_image_changed(self, event=None) -> None:
        del event
        source = self.current_image
        if source is None:
            return
        if self._z_ranges:
            self._clear_z_layers()
        if not self._is_compatible_image_source(source):
            self._set_current_image(None)
            self._refresh_image_layers()
            return
        if self.proofread_store is not None:
            signature_matches = self._image_matches_proof_session(source)
            if not signature_matches:
                # Layer events cannot be vetoed.  Keep the proof state and any
                # size draft, but stop editing/rendering until the original
                # spatial contract is restored or a matching Image is chosen.
                self._proof_detached = True
                self._set_proofreading_off_preserve_draft()
                self.proofreading_toggle.setEnabled(False)
                self._set_proofreading_controls_enabled(False)
                self._remove_roi_layers()
                self._update_info()
                self.update_status(
                    "Image geometry changed; proofreading is paused",
                    "orange",
                )
                return
            if self._proof_detached:
                self._proof_detached = False
                self.update_status(
                    "Image geometry restored; proofreading may be enabled",
                    "green",
                )
        self._remove_roi_layers()
        self._ensure_roi_layers()
        self._refresh_roi_layers()
        self.proofreading_toggle.setEnabled(self._proof_view_allowed())
        self._update_info()

    # Compatibility with the old Z-specific callback name.
    def _on_z_image_changed(self, index: int) -> None:
        self._on_image_changed(index)

    def refresh_z_profile(self) -> None:
        """Count above-threshold pixels per Z for the current Image and time."""

        source = self.z_image_combo.currentData()
        if not isinstance(source, Image):
            return
        time_index = self._z_profile_time_index(source)
        threshold = float(self.z_profile_threshold_spin.value())
        self.z_profile_refresh_btn.setEnabled(False)
        self.z_profile_state_label.setText("computing…")
        try:
            values = z_threshold_count_profile(
                source.data,
                threshold=threshold,
                time_index=time_index,
            )
        except Exception as error:  # noqa: BLE001 - Qt callback boundary
            self._z_profile_source = None
            self._z_profile_time = None
            self.z_profile.clear_profile()
            self.z_profile_state_label.setText("unavailable")
            self.update_status(f"Could not compute Z profile: {error}", "red")
            return
        finally:
            self.z_profile_refresh_btn.setEnabled(True)

        self._z_profile_source = source
        self._z_profile_time = time_index
        self.z_profile.set_profile(values)
        self._sync_z_profile_cuts(self.z_cuts_input.text())
        time_text = f"t={time_index}" if source.ndim == 4 else ""
        self.z_profile_state_label.setText(time_text)

    def _z_profile_time_index(self, source: Image) -> int:
        if source.ndim != 4:
            return 0
        steps = self.viewer.dims.current_step
        value = int(steps[-4]) if len(steps) >= 4 else 0
        return min(max(value, 0), int(source.data.shape[0]) - 1)

    def _sync_z_profile_cuts(self, text: str) -> None:
        try:
            cuts = parse_z_cuts(text)
        except ValueError:
            cuts = ()
        self.z_profile.set_cuts(cuts)

    def _toggle_z_profile_cut(self, cut: int) -> None:
        try:
            cuts = set(parse_z_cuts(self.z_cuts_input.text()))
        except ValueError as error:
            self.update_status(
                f"Fix Z cuts before using the plot: {error}", "orange"
            )
            return
        if cut in cuts:
            cuts.remove(cut)
        else:
            cuts.add(cut)
        source = self.z_image_combo.currentData()
        if not isinstance(source, Image):
            return
        ordered = tuple(sorted(cuts))
        build_z_layer_ranges(int(source.data.shape[-3]), ordered)
        self.z_cuts_input.setText(",".join(str(value) for value in ordered))

    def split_z_layers(self) -> None:
        """Create runtime Image slices and synchronize neuron boxes."""
        try:
            self._create_z_layers()
        except Exception as error:  # noqa: BLE001 - Qt callback boundary
            QMessageBox.critical(
                self,
                "Invalid Z layers",
                f"Could not split the selected layers:\n{error}",
            )
            self.update_status("Z-layer split failed", "red")

    def _create_z_layers(self) -> None:
        source = self.current_image
        if not isinstance(source, Image):
            raise RuntimeError("Select a 3D or 4D Image layer")
        self._validate_image_source(source)

        cuts = parse_z_cuts(self.z_cuts_input.text())
        ranges = build_z_layer_ranges(int(source.data.shape[-3]), cuts)
        if not cuts:
            raise ValueError("Enter at least one Z cut")

        if self._z_ranges:
            self._clear_z_layers()

        self._z_source_image = source
        self._z_ranges = ranges
        self._z_source_image_visible = bool(source.visible)

        try:
            source.events.data.connect(self._on_z_source_image_data_changed)
            for event_name in Z_SOURCE_GEOMETRY_EVENTS:
                getattr(source.events, event_name).connect(
                    self._on_z_source_geometry_changed
                )
            source_index = list(self.viewer.layers).index(source)
            for z_range in ranges:
                derived = self.viewer.add_image(
                    slice_z_range(source.data, z_range),
                    name=f"{source.name} – Layer {z_range.index + 1}",
                    colormap=source.colormap,
                    contrast_limits=tuple(source.contrast_limits),
                    gamma=float(source.gamma),
                    opacity=float(source.opacity),
                    blending="additive",
                    rendering=source.rendering,
                    depiction=source.depiction,
                    attenuation=float(source.attenuation),
                    iso_threshold=float(source.iso_threshold),
                    projection_mode=source.projection_mode,
                    interpolation2d=source.interpolation2d,
                    interpolation3d=source.interpolation3d,
                    rgb=source.rgb,
                    scale=tuple(source.scale),
                    translate=shifted_z_translation(
                        tuple(source.translate),
                        tuple(source.scale),
                        z_range.start,
                    ),
                    axis_labels=tuple(source.axis_labels),
                    units=tuple(source.units),
                    metadata={
                        ROLE_KEY: ROLE_Z_IMAGE,
                        "z_layer_index": z_range.index,
                        "z_start": z_range.start,
                        "z_stop": z_range.stop,
                        "z_session": self._z_session_token,
                    },
                )
                self._z_image_layers.append(derived)
                current_index = list(self.viewer.layers).index(derived)
                target_index = source_index + len(self._z_image_layers)
                if current_index != target_index:
                    self.viewer.layers.move(current_index, target_index)

            source.visible = False

            self.z_view_combo.blockSignals(True)
            self.z_view_combo.clear()
            self.z_view_combo.addItem("All", None)
            for z_range in ranges:
                self.z_view_combo.addItem(
                    f"Layer {z_range.index + 1} · "
                    f"z {z_range.start}–{z_range.stop - 1}",
                    z_range.index,
                )
            self.z_view_combo.setCurrentIndex(0)
            self.z_view_combo.blockSignals(False)
            self.z_view_combo.setEnabled(True)
            self.clear_z_btn.setEnabled(True)
            self.z_image_combo.setEnabled(False)
            self._z_active_index = None
            self._apply_z_view()
        except Exception:
            self._clear_z_layers()
            raise

        self.update_status(
            f"Split {source.name} into {len(ranges)} Z layers",
            "green",
        )

    def _validate_image_source(self, source: Image) -> None:
        if source.rgb:
            raise ValueError("RGB Image layers are not supported")
        if source.multiscale:
            raise ValueError("Multiscale Image layers are not supported")
        if source.depiction != "volume":
            raise ValueError("Only volume depiction is supported")
        if len(source.experimental_clipping_planes):
            raise ValueError(
                "Experimental clipping planes are not supported"
            )
        if source.ndim not in (3, 4):
            raise ValueError("Image must use (z,y,x) or (t,z,y,x)")
        if not self._uses_axis_aligned_transform(source):
            raise ValueError("Image transform must be axis-aligned")
        data_module = type(source.data).__module__.split(".", 1)[0]
        if data_module == "zarr":
            raise ValueError(
                "Direct Zarr arrays are not supported; wrap the array "
                "as a Dask array before use"
            )

    def _is_compatible_image_source(self, source: Image) -> bool:
        try:
            self._validate_image_source(source)
        except (TypeError, ValueError):
            return False
        return True

    @staticmethod
    def _uses_axis_aligned_transform(layer: Image) -> bool:
        ndim = layer.ndim
        return (
            np.allclose(layer.rotate, np.eye(ndim))
            and np.allclose(layer.shear, 0.0)
            and np.allclose(layer.affine.affine_matrix, np.eye(ndim + 1))
        )

    def _on_z_source_image_data_changed(self, event=None) -> None:
        del event
        if self._z_ranges:
            self._clear_z_layers()
            self.update_status(
                "Z layers cleared because the source Image changed",
                "orange",
            )

    def _on_z_source_geometry_changed(self, event=None) -> None:
        del event
        if self._z_ranges:
            self._clear_z_layers()
            self.update_status(
                "Z layers cleared because a source transform changed",
                "orange",
            )

    def _on_z_view_changed(self, index: int) -> None:
        del index
        if not self._z_ranges:
            return
        value = self.z_view_combo.currentData()
        previous = self._z_active_index
        target = None if value is None else int(value)
        if (
            target != previous
            and self._proof_size_draft_dirty
            and not self._confirm_proof_size_draft("changing Z view")
        ):
            self.z_view_combo.blockSignals(True)
            try:
                restore = self.z_view_combo.findData(previous)
                self.z_view_combo.setCurrentIndex(restore)
            finally:
                self.z_view_combo.blockSignals(False)
            return
        self._z_active_index = target
        self._apply_z_view()

    def _active_z_range(self) -> ZLayerRange | None:
        if self._z_active_index is None:
            return None
        if not 0 <= self._z_active_index < len(self._z_ranges):
            return None
        return self._z_ranges[self._z_active_index]

    def _apply_z_view(self) -> None:
        if not self._z_ranges:
            return
        active_range = self._active_z_range()
        for z_range, layer in zip(
            self._z_ranges, self._z_image_layers, strict=True
        ):
            layer.visible = (
                active_range is None or z_range.index == active_range.index
            )

        if self._z_source_image is not None:
            self._z_source_image.visible = False
        if active_range is not None:
            if self.proofreading_enabled:
                self._proof_cancel_or_exit()
            self._clamp_z_to_active_layer(active_range)

        self._refresh_selection_item_styles()
        self._refresh_roi_layers()
        self._update_info()

    def _clamp_z_to_active_layer(self, z_range: ZLayerRange) -> None:
        if self.viewer.dims.ndisplay != 2:
            return
        steps = list(self.viewer.dims.current_step)
        if len(steps) < 3:
            return
        z_axis = len(steps) - 3
        current_z = int(steps[z_axis])
        if not z_range.start <= current_z < z_range.stop:
            steps[z_axis] = (z_range.start + z_range.stop - 1) // 2
            self.viewer.dims.current_step = tuple(steps)

    def clear_z_layers(self) -> None:
        """Remove runtime Z layers and restore source visibility."""
        if not self._z_ranges:
            return
        self._clear_z_layers()
        self.update_status("Z layers cleared", "green")

    def _clear_z_layers(self) -> None:
        source_image = self._z_source_image
        image_visible = self._z_source_image_visible

        if source_image is not None:
            with suppress(TypeError, ValueError):
                source_image.events.data.disconnect(
                    self._on_z_source_image_data_changed
                )
            for event_name in Z_SOURCE_GEOMETRY_EVENTS:
                with suppress(TypeError, ValueError):
                    getattr(source_image.events, event_name).disconnect(
                        self._on_z_source_geometry_changed
                    )
        self._z_cleanup = True
        try:
            managed = [
                layer
                for layer in self.viewer.layers
                if layer in self._z_image_layers
                or (
                    layer.metadata.get(ROLE_KEY) in MANAGED_Z_ROLES
                    and layer.metadata.get("z_session")
                    == self._z_session_token
                )
            ]
            for layer in managed:
                if layer in self.viewer.layers:
                    self.viewer.layers.remove(layer)
        finally:
            self._z_cleanup = False

        if (
            source_image is not None
            and source_image in self.viewer.layers
            and image_visible is not None
        ):
            source_image.visible = image_visible
        self._z_ranges = ()
        self._z_source_image = None
        self._z_image_layers.clear()
        self._z_active_index = None
        self._z_source_image_visible = None

        self.z_view_combo.blockSignals(True)
        self.z_view_combo.clear()
        self.z_view_combo.addItem("All", None)
        self.z_view_combo.blockSignals(False)
        self.z_view_combo.setEnabled(False)
        self.clear_z_btn.setEnabled(False)
        self._refresh_image_layers()
        self._refresh_selection_item_styles()
        self._refresh_roi_layers()
        self._update_info()

    def _on_layer_removed(self, event) -> None:
        if self._closed:
            return
        removed = getattr(event, "value", None)
        metadata = getattr(removed, "metadata", {})
        role = metadata.get(ROLE_KEY)
        if role in MANAGED_Z_ROLES:
            if (
                metadata.get("z_session") == self._z_session_token
                and not self._z_cleanup
            ):
                self._clear_z_layers()
            return
        if (
            isinstance(removed, Points | Vectors)
            and removed.metadata.get(ROLE_KEY) in MANAGED_ROI_ROLES
        ):
            if role == ROLE_PROOF_TARGET:
                self._clear_proof_target()
            return
        if removed is self._z_source_image:
            self._clear_z_layers()
        if removed is self.current_image:
            # napari has already removed the layer, so this transition cannot
            # be vetoed.  Preserve dirty proof state in a detached session.
            self._set_current_image(None, force=True)
            if self._proof_session_has_content():
                self._proof_detached = True
                self.update_status(
                    "Image removed; proofreading state is retained",
                    "orange",
                )
        self._refresh_image_layers()

    def closeEvent(self, event: QCloseEvent) -> None:
        if not self.shutdown():
            event.ignore()
            return
        event.accept()

    def shutdown(self, *, force: bool = False) -> bool:
        """Restore managed display state and disconnect all external events."""
        if self._closed:
            return True
        if not force and not self._confirm_proof_transition("closing the widget"):
            return False
        self._remove_proof_key_filter()
        self._remove_proof_mouse_callback()
        self._clear_proof_target()
        self.proofreading_enabled = False
        self._closed = True
        self._restore_orientation_baseline()
        self._disconnect_viewer_events()
        self._unbind_keys()
        self._clear_z_layers()
        self._disconnect_current_image_events()
        self.current_image = None
        self._remove_roi_layers()
        return True

    # ------------------------------------------------------------------
    # ROI identities, checkable selection, and navigation
    # ------------------------------------------------------------------
    def _refresh_available_ids(self, *, select_first: bool) -> None:
        ids = (
            list(self.proofread_store.neuron_ids)
            if self.proofread_store is not None
            else self.roi_dataset.neuron_ids
            if self.roi_dataset is not None
            else []
        )
        self._available_ids = ids

        if self.active_id not in ids:
            self.active_id = None
        self.checked_ids.intersection_update(ids)

        if select_first and ids and self.active_id is None:
            valid = self._valid_navigation_ids()
            self.active_id = valid[0] if valid else ids[0]
            self.checked_ids = {self.active_id}

        self._rebuild_selection_items()
        self._sync_annotation_ids()

    def _rebuild_selection_items(self) -> None:
        self._ui_sync = True
        try:
            self.selection_tree.clear()
            self._selection_items.clear()
            for display_id in self._available_ids:
                item = QTreeWidgetItem()
                item.setFlags(
                    item.flags()
                    | Qt.ItemIsUserCheckable
                    | Qt.ItemIsSelectable
                    | Qt.ItemIsEnabled
                )
                item.setData(0, Qt.UserRole, display_id)
                item.setCheckState(
                    0,
                    Qt.Checked
                    if display_id in self.checked_ids
                    else Qt.Unchecked,
                )
                self.selection_tree.addTopLevelItem(item)
                self._selection_items[display_id] = item
        finally:
            self._ui_sync = False
        self._refresh_selection_item_styles()

    def _refresh_selection_item_styles(self) -> None:
        names = self._annotation_names()
        valid_time_ids = set(self._valid_time_ids())
        self._ui_sync = True
        try:
            for display_id in self._available_ids:
                self._style_selection_item(
                    display_id,
                    valid_at_time=display_id in valid_time_ids,
                    in_active_layer=self._id_in_active_z_layer(display_id),
                    biological_name=names.get(display_id, ""),
                )
            if self.active_id is None:
                self.selection_tree.clearSelection()
                self.selection_tree.setCurrentItem(None)
            else:
                item = self._selection_items.get(self.active_id)
                if item is not None:
                    self.selection_tree.setCurrentItem(item)
                    self.selection_tree.scrollToItem(item)
        finally:
            self._ui_sync = False

    def _style_selection_item(
        self,
        display_id: int,
        *,
        valid_at_time: bool | None = None,
        in_active_layer: bool | None = None,
        biological_name: str | None = None,
    ) -> None:
        item = self._selection_items.get(display_id)
        if item is None:
            return
        if valid_at_time is None:
            valid_at_time = display_id in set(self._valid_time_ids())
        if in_active_layer is None:
            in_active_layer = self._id_in_active_z_layer(display_id)
        if biological_name is None:
            biological_name = self._annotation_names().get(display_id, "")
        biological_name = biological_name.strip()

        text = f"Neuron {display_id}"
        if biological_name:
            text += f" · {biological_name}"
        if self.roi_dataset is not None and not valid_at_time:
            text += " (missing)"
        item.setText(1, text)
        item.setCheckState(
            0,
            Qt.Checked if display_id in self.checked_ids else Qt.Unchecked,
        )
        font = item.font(1)
        font.setBold(display_id == self.active_id)
        item.setFont(1, font)
        if valid_at_time and in_active_layer:
            rgba = self._display_color(display_id)
            color = QColor.fromRgbF(*rgba[:3])
        else:
            color = QColor("#777777")
        item.setForeground(1, QBrush(color))
        if display_id in self._search_match_ids:
            background = self.selection_tree.palette().color(
                QPalette.Highlight
            )
            background.setAlpha(70)
            brush = QBrush(background)
        else:
            brush = QBrush()
        item.setBackground(0, brush)
        item.setBackground(1, brush)

    def _on_search_text_changed(self, text: str) -> None:
        del text
        self._recompute_search_matches()

    def _recompute_search_matches(self) -> None:
        self._search_match_ids = _match_neuron_ids(
            self._available_ids,
            self._annotation_names(),
            self.neuron_search_input.text(),
        )
        self._search_cursor = -1
        self._update_search_controls()
        self._refresh_selection_item_styles()

    def _update_search_controls(self) -> None:
        match_count = len(self._search_match_ids)
        if not self.neuron_search_input.text().strip():
            text = ""
        elif self._search_cursor >= 0 and match_count:
            text = f"{self._search_cursor + 1}/{match_count} matches"
        else:
            text = f"{match_count} matches"
        self.search_matches_label.setText(text)
        self.check_matches_btn.setEnabled(bool(self._search_match_ids))

    def _reset_search(self) -> None:
        self.neuron_search_input.blockSignals(True)
        try:
            self.neuron_search_input.clear()
        finally:
            self.neuron_search_input.blockSignals(False)
        self._search_match_ids = []
        self._search_cursor = -1
        self._update_search_controls()

    def _activate_next_search_match(self) -> None:
        if not self._search_match_ids:
            if self.neuron_search_input.text().strip():
                self.update_status("No neuron matches the search", "orange")
            return
        self._search_cursor = (
            self._search_cursor + 1
        ) % len(self._search_match_ids)
        neuron_id = self._search_match_ids[self._search_cursor]
        self._update_search_controls()
        self.activate_id(neuron_id, locate=True)

    def check_search_matches(self) -> None:
        """Add every search result to checked IDs without changing active."""
        if not self._search_match_ids:
            return
        self.checked_ids.update(self._search_match_ids)
        self._selection_changed(locate=False)

    def _display_color(self, display_id: int) -> np.ndarray:
        return np.asarray(neuron_color(display_id), dtype=float)

    def _on_selection_item_changed(
        self, item: QTreeWidgetItem, column: int
    ) -> None:
        if self._ui_sync or column != 0:
            return
        display_id = int(item.data(0, Qt.UserRole))
        checked = item.checkState(0) == Qt.Checked
        changes_active = (checked and display_id != self.active_id) or (
            not checked and display_id == self.active_id
        )
        if (
            changes_active
            and self._proof_size_draft_dirty
            and not self._confirm_proof_size_draft("changing active neuron")
        ):
            self._ui_sync = True
            try:
                item.setCheckState(
                    0,
                    Qt.Checked
                    if display_id in self.checked_ids
                    else Qt.Unchecked,
                )
            finally:
                self._ui_sync = False
            self._refresh_selection_item_styles()
            return
        if checked:
            self.checked_ids.add(display_id)
            self.active_id = display_id
        else:
            self.checked_ids.discard(display_id)
            if self.active_id == display_id:
                self.active_id = None
        self._selection_changed(locate=checked)

    def _on_selection_item_clicked(
        self, item: QTreeWidgetItem, column: int
    ) -> None:
        if self._ui_sync or column == 0:
            return
        display_id = int(item.data(0, Qt.UserRole))
        self.activate_id(display_id, locate=True)

    def activate_id(self, display_id: int, *, locate: bool = True) -> None:
        """Check and activate one available identity."""
        if display_id not in self._available_ids:
            return
        if (
            display_id != self.active_id
            and self._proof_size_draft_dirty
            and not self._confirm_proof_size_draft("changing active neuron")
        ):
            self._refresh_selection_item_styles()
            self._sync_annotation_to_active()
            return
        self.checked_ids.add(display_id)
        self.active_id = display_id
        self._selection_changed(locate=locate)

    def check_all(self) -> None:
        """Check every available identity without changing an existing active."""
        self.checked_ids = set(self._available_ids)
        locate = False
        if self.active_id is None:
            valid = self._valid_navigation_ids()
            if valid:
                self.active_id = valid[0]
                locate = True
        self._selection_changed(locate=locate)

    def check_none(self) -> None:
        """Clear both the checked collection and active identity."""
        if (
            self.active_id is not None
            and self._proof_size_draft_dirty
            and not self._confirm_proof_size_draft("clearing selection")
        ):
            return
        self.checked_ids.clear()
        self.active_id = None
        self._selection_changed(locate=False)

    def _selection_changed(self, *, locate: bool) -> None:
        draft_target = self._proof_size_draft_target
        if (
            self._proof_size_draft_dirty
            and draft_target is not None
            and self.active_id != draft_target[1]
            and not self._confirm_proof_size_draft("changing active neuron")
        ):
            # Some Qt selection signals update active_id before reaching this
            # common path. Restore the draft owner when the user cancels.
            self.active_id = draft_target[1]
            self.checked_ids.add(draft_target[1])
        self._refresh_selection_item_styles()
        self._refresh_roi_layers()
        self._update_proof_size_controls()
        self._sync_annotation_to_active()
        if locate:
            self._locate_active_box()
        self._update_info()

    def _valid_navigation_ids(self) -> list[int]:
        return [
            display_id
            for display_id in self._valid_time_ids()
            if self._id_in_active_z_layer(display_id)
        ]

    def _valid_time_ids(self) -> list[int]:
        if self.roi_dataset is None:
            return list(self._available_ids)
        volume_index = self._current_volume_index()
        if self.proofread_store is not None:
            return (
                self.proofread_store.valid_ids_at_volume_index(volume_index)
                if volume_index is not None
                else []
            )
        return self.roi_dataset.valid_ids(self._viewer_time())

    def _id_in_active_z_layer(self, display_id: int) -> bool:
        active_range = self._active_z_range()
        if active_range is None or self.roi_dataset is None:
            return True
        volume_index = self._current_volume_index()
        box = (
            self.proofread_store.resolve(volume_index, display_id)
            if self.proofread_store is not None and volume_index is not None
            else self.roi_dataset.get_box(self._viewer_time(), display_id)
        )
        if box is None:
            return False
        owner = find_z_layer(box.center_zyx[0], self._z_ranges)
        return owner is not None and owner.index == active_range.index

    def navigate(self, step: int) -> None:
        valid_ids = self._valid_navigation_ids()
        if not valid_ids:
            self.update_status(
                "No valid neuron at the current volume", "orange"
            )
            return
        if self.active_id not in valid_ids:
            next_id = valid_ids[0] if step >= 0 else valid_ids[-1]
        else:
            index = valid_ids.index(self.active_id)
            next_id = valid_ids[(index + step) % len(valid_ids)]
        self.activate_id(next_id, locate=True)

    def navigate_checked(self, step: int) -> None:
        """Activate the adjacent checked ID that is currently navigable."""
        candidates = [
            neuron_id
            for neuron_id in self._valid_navigation_ids()
            if neuron_id in self.checked_ids
        ]
        if not candidates:
            self.update_status(
                "No checked neuron is navigable in the current time/Z view",
                "orange",
            )
            return

        if self.active_id is None:
            next_id = candidates[0] if step >= 0 else candidates[-1]
        elif self.active_id in candidates:
            index = candidates.index(self.active_id)
            next_id = candidates[(index + step) % len(candidates)]
        elif step >= 0:
            next_id = next(
                (item_id for item_id in candidates if item_id > self.active_id),
                candidates[0],
            )
        else:
            next_id = next(
                (
                    item_id
                    for item_id in reversed(candidates)
                    if item_id < self.active_id
                ),
                candidates[-1],
            )
        self.activate_id(next_id, locate=True)

    def _previous_key(self, viewer=None) -> None:
        del viewer
        self.navigate(-1)

    def _next_key(self, viewer=None) -> None:
        del viewer
        self.navigate(1)

    def _previous_checked_key(self, viewer=None) -> None:
        del viewer
        self.navigate_checked(-1)

    def _next_checked_key(self, viewer=None) -> None:
        del viewer
        self.navigate_checked(1)

    def _step_z(self, step: int) -> None:
        if self.current_image is None or self.viewer.dims.ndisplay != 2:
            return
        steps = self.viewer.dims.current_step
        if len(steps) < 3:
            return
        z_axis = len(steps) - 3
        z_start = 0
        z_stop = self._shape_zyx()[0] - 1
        active_range = self._active_z_range()
        if active_range is not None:
            z_start = active_range.start
            z_stop = active_range.stop - 1
        target = int(np.clip(steps[z_axis] + step, z_start, z_stop))
        self.viewer.dims.set_current_step(z_axis, target)

    def _previous_z_key(self, viewer=None) -> None:
        del viewer
        self._step_z(-1)

    def _next_z_key(self, viewer=None) -> None:
        del viewer
        self._step_z(1)

    def _step_time(self, step: int) -> None:
        if self.current_image is None or self.current_image.ndim != 4:
            return
        steps = self.viewer.dims.current_step
        if len(steps) < 4:
            return
        time_axis = len(steps) - 4
        time_stop = int(self.current_image.data.shape[0]) - 1
        target = int(np.clip(steps[time_axis] + step, 0, time_stop))
        if (
            target != int(steps[time_axis])
            and self._proof_size_draft_dirty
            and not self._confirm_proof_size_draft("changing volume")
        ):
            return
        self.viewer.dims.set_current_step(time_axis, target)

    def _previous_time_key(self, viewer=None) -> None:
        del viewer
        self._step_time(-1)

    def _next_time_key(self, viewer=None) -> None:
        del viewer
        self._step_time(1)

    def _previous_time_fast_key(self, viewer=None) -> None:
        del viewer
        self._step_time(-10)

    def _next_time_fast_key(self, viewer=None) -> None:
        del viewer
        self._step_time(10)

    # ------------------------------------------------------------------
    # ROI loading and geometry layers
    # ------------------------------------------------------------------
    def load_roi_npy(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Load neuron_pt_tuple",
            "",
            "NumPy arrays (*.npy);;All files (*)",
        )
        if not path:
            return

        try:
            self.load_roi_path(path)
        except (OSError, RuntimeError, TypeError, ValueError) as error:
            QMessageBox.critical(
                self, "Invalid ROI file", f"Could not load ROI data:\n{error}"
            )
            self.update_status("ROI load failed", "red")

    def load_roi_path(self, path: str | Path) -> None:
        """Load and activate a read-only ROI NPY without opening a dialog."""
        if self.current_image is None:
            raise RuntimeError("Select an Image layer before loading ROI data")
        if self.current_image.ndim not in (3, 4):
            raise ValueError(
                "ROI overlays require a (z,y,x) or (t,z,y,x) Image layer"
            )

        dataset = NeuronBoxDataset.from_npy(
            path,
            z_divisor=self.z_divisor_spin.value(),
            volume_start=self.volume_start_spin.value(),
            volume_stride=self.volume_stride_spin.value(),
        )
        source_path = Path(path)

        if self.roi_dataset is not None and not self._confirm_proof_transition(
            "loading another ROI"
        ):
            return
        if self.proofreading_enabled:
            self._proof_cancel_or_exit()
        else:
            self._remove_proof_mouse_callback()
            self._clear_proof_target()

        self.roi_dataset = dataset
        self.proofread_store = (
            ProofreadStore(
                dataset,
                image_signature=self._proof_image_signature(
                    self.current_image
                ),
            )
            if ProofreadStore
            else None
        )
        self._proof_detached = False
        self.roi_path_input.setText(str(source_path))
        self.unload_roi_btn.setEnabled(True)
        self._set_roi_config_enabled(False)
        self.active_id = None
        self.checked_ids.clear()
        self._reset_search()
        self._refresh_available_ids(select_first=True)
        self._ensure_roi_layers()
        self._refresh_roi_layers()
        self._update_roi_info()
        self._update_proof_current_box_status()
        self.proofreading_toggle.setEnabled(self._proof_view_allowed())
        self._set_proofreading_controls_enabled(False)
        self.update_status(f"Loaded ROI: {source_path.name}", "green")

    def unload_roi(self, *, force: bool = False) -> bool:
        if not force and not self._confirm_proof_transition("unloading ROI"):
            return False
        if self.proofreading_enabled:
            self._proof_cancel_or_exit()
        else:
            self._remove_proof_mouse_callback()
            self._clear_proof_target()
        self.proofread_store = None
        self.roi_dataset = None
        self.roi_path_input.clear()
        self.unload_roi_btn.setEnabled(False)
        self._set_roi_config_enabled(True)
        self._remove_roi_layers()
        self.active_id = None
        self.checked_ids.clear()
        self._available_ids = []
        self._reset_search()
        self.roi_info_label.setText("No ROI loaded")
        self.proofreading_toggle.setEnabled(False)
        self._set_proofreading_controls_enabled(False)
        self._rebuild_selection_items()
        self._update_proof_current_box_status()
        self._update_info()
        self.update_status("ROI unloaded", "green")
        return True

    def _set_roi_config_enabled(self, enabled: bool) -> None:
        self.z_divisor_spin.setEnabled(enabled)
        self.volume_start_spin.setEnabled(enabled)
        self.volume_stride_spin.setEnabled(enabled)
        self.load_roi_btn.setEnabled(enabled and self.current_image is not None)

    def _update_roi_info(self) -> None:
        if self.roi_dataset is None:
            return
        viewer_t = self._viewer_time()
        volume_index = self._current_volume_index()
        valid = len(self._valid_time_ids())
        volume_text = "out of range" if volume_index is None else str(volume_index)
        self.roi_info_label.setText(
            f"T={self.roi_dataset.time_count}, "
            f"N={self.roi_dataset.neuron_count}; "
            f"Image t={viewer_t} → volume={volume_text}; "
            f"{valid} valid boxes"
        )

    def _viewer_time(self) -> int:
        if self.current_image is None or self.current_image.ndim == 3:
            return 0
        steps = self.viewer.dims.current_step
        return int(steps[-4]) if len(steps) >= 4 else 0

    def _viewer_z(self) -> int:
        if self.current_image is None:
            return 0
        steps = self.viewer.dims.current_step
        if self.current_image.ndim == 4 and len(steps) >= 4:
            return int(steps[-3])
        if len(steps) >= 3:
            return int(steps[-3])
        return 0

    def _shape_zyx(self) -> tuple[int, int, int]:
        if self.current_image is None:
            raise RuntimeError("No Image layer selected")
        shape = self.current_image.data.shape
        return tuple(int(value) for value in shape[-3:])

    def _view_axes_supported(self) -> bool:
        if self.current_image is None:
            return False
        ndim = int(self.viewer.dims.ndim)
        order = tuple(self.viewer.dims.order)
        y_axis = ndim - 2
        x_axis = ndim - 1
        if self.viewer.dims.ndisplay == 2:
            return set(order[-2:]) == {y_axis, x_axis}
        return (
            order[-3] == ndim - 3
            and set(order[-2:]) == {y_axis, x_axis}
        )

    def _on_dims_changed(self, event=None) -> None:
        del event
        if self._closed:
            return
        viewer_time = self._viewer_time()
        time_changed = bool(
            self.current_image is not None
            and self.current_image.ndim == 4
            and self._last_viewer_time is not None
            and viewer_time != self._last_viewer_time
        )
        self._last_viewer_time = (
            viewer_time
            if self.current_image is not None and self.current_image.ndim == 4
            else None
        )
        current_volume = self._current_volume_index()
        current_target_context = (
            current_volume,
            viewer_time,
            self._viewer_z(),
        )
        if (
            self._proof_target_context is not None
            and self._proof_target_context != current_target_context
        ):
            self._clear_proof_target()
        draft_target = self._proof_size_draft_target
        target_changed = bool(
            self._proof_size_draft_dirty
            and draft_target is not None
            and current_volume != draft_target[0]
        )
        view_changed = self.proofreading_enabled and not self._proof_view_allowed()
        if target_changed or view_changed:
            allowed = self._confirm_proof_size_draft(
                "changing volume or display mode"
            )
            if view_changed or not allowed:
                # An external napari dims change has already happened and
                # cannot reliably be vetoed. Stop proofreading; on Cancel the
                # draft remains attached to its original target.
                if allowed:
                    self.proofreading_toggle.blockSignals(True)
                    self.proofreading_toggle.setChecked(False)
                    self.proofreading_toggle.blockSignals(False)
                    self._set_proofreading_enabled(False)
                else:
                    self._set_proofreading_off_preserve_draft()
                if not allowed:
                    self.update_status(
                        "Proofreading paused; size draft retained", "orange"
                    )
        if (
            self._orientation_baseline_ndim is not None
            and self._orientation_baseline_ndim != int(self.viewer.dims.ndim)
            and not self._orientation_applying
        ):
            self._rebase_orientation()
        if (
            self._z_profile_source is not None
            and self._z_profile_source.ndim == 4
            and self._z_profile_time is not None
            and self._z_profile_time_index(self._z_profile_source)
            != self._z_profile_time
        ):
            self.z_profile_state_label.setText(
                f"t={self._z_profile_time} · Refresh"
            )
        if (
            self.roi_dataset is not None
            and self.current_image is not None
            and self.current_image in self.viewer.layers
        ):
            self._update_roi_info()
            self._refresh_selection_item_styles()
            self._refresh_roi_layers()
            self._update_proof_size_controls()
            if time_changed:
                self._locate_active_box()
        else:
            # Keep the empty-state label accurate while an Image/ROI
            # transition temporarily leaves no resolvable observation.
            self._update_proof_current_box_status()
        self._update_orientation_controls_enabled()
        self.proofreading_toggle.setEnabled(self._proof_view_allowed())
        self._set_proofreading_controls_enabled(
            self.proofreading_enabled and self._proof_view_allowed()
        )

    def _ensure_roi_layers(self) -> None:
        if (
            self.current_image is None
            or self.roi_dataset is None
            or self.current_image.ndim not in (3, 4)
        ):
            return
        # ``Viewer.add_*`` selects the newly-created layer.  Preserve the
        # user's active layer while constructing runtime ROI overlays; without
        # this, loading an ROI leaves the transparent label overlay active and
        # Proofreading immediately rejects the toggle because the source Image
        # is no longer active.
        previous_active = self.viewer.layers.selection.active
        ndim = self.current_image.ndim
        empty_vectors = np.empty((0, 2, ndim), dtype=float)
        empty_points = np.empty((0, ndim), dtype=float)
        axis_labels = tuple(self.current_image.axis_labels)
        units = tuple(self.current_image.units)

        legacy_layer = self._managed_vector_layer(LEGACY_ROLE_ALL)
        if legacy_layer is not None and legacy_layer in self.viewer.layers:
            self.viewer.layers.remove(legacy_layer)

        if self._managed_vector_layer(ROLE_SELECTED) is None:
            self.viewer.add_vectors(
                empty_vectors,
                ndim=ndim,
                name="Neuron boxes – selected",
                vector_style="line",
                edge_width=1.0,
                edge_color="white",
                opacity=0.25,
                blending="translucent",
                scale=tuple(self.current_image.scale),
                translate=tuple(self.current_image.translate),
                axis_labels=axis_labels,
                units=units,
                metadata={ROLE_KEY: ROLE_SELECTED},
            )
        if self._managed_vector_layer(ROLE_ACTIVE) is None:
            self.viewer.add_vectors(
                empty_vectors,
                ndim=ndim,
                name="Neuron box – active",
                vector_style="line",
                edge_width=2.0,
                edge_color="yellow",
                opacity=1.0,
                blending="translucent",
                scale=tuple(self.current_image.scale),
                translate=tuple(self.current_image.translate),
                axis_labels=axis_labels,
                units=units,
                metadata={ROLE_KEY: ROLE_ACTIVE},
            )
        if self._managed_box_label_layer() is None:
            box_label_layer = self.viewer.add_points(
                empty_points,
                ndim=ndim,
                name="Neuron labels – selected",
                size=1,
                face_color="transparent",
                border_color="transparent",
                border_width=0,
                opacity=1.0,
                blending="translucent",
                features={
                    "neuron_id": np.empty(0, dtype=int),
                    "display_text": np.empty(0, dtype=str),
                },
                text={
                    "string": "{display_text}",
                    "color": self._box_label_color,
                    "size": 12,
                    "anchor": "center",
                },
                visible=self.show_box_labels_checkbox.isChecked(),
                scale=tuple(self.current_image.scale),
                translate=tuple(self.current_image.translate),
                axis_labels=axis_labels,
                units=units,
                metadata={ROLE_KEY: ROLE_BOX_LABELS},
            )
            box_label_layer.editable = False
        if previous_active is not None and previous_active in self.viewer.layers:
            self.viewer.layers.selection.active = previous_active
        elif previous_active is None:
            self.viewer.layers.selection.active = None
        elif self.current_image in self.viewer.layers:
            # The previous active layer may have been the legacy managed layer
            # removed above.  The source Image is the safe fallback.
            self.viewer.layers.selection.active = self.current_image
    def _managed_vector_layer(self, role: str) -> Vectors | None:
        for layer in self.viewer.layers:
            if (
                isinstance(layer, Vectors)
                and layer.metadata.get(ROLE_KEY) == role
            ):
                return layer
        return None

    def _managed_box_label_layer(self) -> Points | None:
        for layer in self.viewer.layers:
            if (
                isinstance(layer, Points)
                and layer.metadata.get(ROLE_KEY) == ROLE_BOX_LABELS
            ):
                return layer
        return None

    def _ensure_proof_target_layer(self) -> Vectors | None:
        """Create the session-only, read-only proofreading crosshair layer."""
        if self.current_image is None or self.current_image.ndim not in (3, 4):
            return None
        active = self.viewer.layers.selection.active
        # Releases before the vector crosshair used a Points marker under the
        # same role.  Remove that legacy layer rather than returning it, so a
        # reopened widget always gets independently tunable line geometry.
        legacy_points = [
            layer
            for layer in self.viewer.layers
            if isinstance(layer, Points)
            and layer.metadata.get(ROLE_KEY) == ROLE_PROOF_TARGET
        ]
        for legacy in legacy_points:
            if legacy in self.viewer.layers:
                self.viewer.layers.remove(legacy)

        layer = self._managed_proof_target_layer()
        if layer is not None:
            layer.editable = False
            if active in self.viewer.layers:
                self.viewer.layers.selection.active = active
            return layer
        layer = self.viewer.add_vectors(
            np.empty((0, 2, self.current_image.ndim), dtype=float),
            ndim=self.current_image.ndim,
            name="Proofreading target",
            vector_style="line",
            edge_color="cyan",
            edge_width=float(PROOF_TARGET_EDGE_WIDTH),
            opacity=1.0,
            blending="translucent",
            scale=tuple(self.current_image.scale),
            translate=tuple(self.current_image.translate),
            axis_labels=tuple(self.current_image.axis_labels),
            units=tuple(self.current_image.units),
            visible=False,
            metadata={ROLE_KEY: ROLE_PROOF_TARGET},
        )
        layer.editable = False
        self.viewer.layers.selection.active = (
            active if active in self.viewer.layers else None
        )
        return layer

    def _managed_proof_target_layer(self) -> Vectors | None:
        for layer in self.viewer.layers:
            if (
                isinstance(layer, Vectors)
                and layer.metadata.get(ROLE_KEY) == ROLE_PROOF_TARGET
            ):
                return layer
        return None

    def _remove_roi_layers(self) -> None:
        self._clear_proof_target()
        managed = [
            layer
            for layer in self.viewer.layers
            if isinstance(layer, Points | Vectors)
            and layer.metadata.get(ROLE_KEY) in MANAGED_ROI_ROLES
        ]
        for layer in managed:
            if layer in self.viewer.layers:
                self.viewer.layers.remove(layer)

    def _refresh_roi_layers(self, event=None) -> None:
        del event
        if (
            self.roi_dataset is None
            or self.current_image is None
            or self.current_image.ndim not in (3, 4)
        ):
            return
        selected_layer = self._managed_vector_layer(ROLE_SELECTED)
        active_layer = self._managed_vector_layer(ROLE_ACTIVE)
        box_label_layer = self._managed_box_label_layer()
        if selected_layer is None or active_layer is None:
            return

        if not self._view_axes_supported():
            self._set_vector_data(selected_layer, [], [])
            self._set_vector_data(active_layer, [], [], active=True)
            if box_label_layer is not None:
                self._set_box_label_data(box_label_layer, [], [], [])
            self.update_status(
                "ROI overlays require spatial y/x axes in 2D or z/y/x axes in 3D",
                "orange",
            )
            return

        viewer_t = self._viewer_time()
        volume_index = self._current_volume_index()
        z_index = self._viewer_z()
        shape_zyx = self._shape_zyx()
        selected_vectors: list[np.ndarray] = []
        selected_neuron_ids: list[int] = []
        active_vectors: list[np.ndarray] = []
        active_ids: list[int] = []
        label_points: list[np.ndarray] = []
        label_neuron_ids: list[int] = []
        display_texts: list[str] = []
        biological_names = self._annotation_names()
        show_box_labels = (
            box_label_layer is not None
            and self.show_box_labels_checkbox.isChecked()
        )

        valid_ids = (
            self.proofread_store.valid_ids_at_volume_index(volume_index)
            if self.proofread_store is not None and volume_index is not None
            else self.roi_dataset.valid_ids(viewer_t)
        )
        for neuron_id in valid_ids:
            if (
                neuron_id not in self.checked_ids
                and neuron_id != self.active_id
            ):
                continue
            box = (
                self.proofread_store.resolve(volume_index, neuron_id)
                if self.proofread_store is not None and volume_index is not None
                else self.roi_dataset.get_box(viewer_t, neuron_id)
            )
            if box is None:
                continue
            active_range = self._active_z_range()
            if active_range is not None:
                owner = find_z_layer(box.center_zyx[0], self._z_ranges)
                if owner is None or owner.index != active_range.index:
                    continue
            if self.viewer.dims.ndisplay == 2:
                geometry = box_vectors_2d(box, z_index, shape_zyx=shape_zyx)
                label_point = box_label_point_2d(
                    box, z_index, shape_zyx=shape_zyx
                )
            else:
                geometry = box_vectors_3d(box, shape_zyx=shape_zyx)
                label_point = box_label_point_3d(box, shape_zyx=shape_zyx)
            if self.current_image.ndim == 4:
                geometry = add_time_axis(geometry, viewer_t)
                if label_point is not None:
                    label_point = np.concatenate(
                        ([float(viewer_t)], label_point)
                    )
            if len(geometry):
                if neuron_id in self.checked_ids:
                    selected_vectors.append(geometry)
                    selected_neuron_ids.extend([neuron_id] * len(geometry))
                    if show_box_labels and label_point is not None:
                        label_points.append(
                            np.asarray(label_point, dtype=float)
                        )
                        label_neuron_ids.append(neuron_id)
                        biological = biological_names.get(neuron_id, "")
                        display_texts.append(
                            self._box_label_text(neuron_id, biological)
                        )
                if neuron_id == self.active_id:
                    active_vectors.append(geometry)
                    active_ids.extend([neuron_id] * len(geometry))

        self._set_vector_data(
            selected_layer, selected_vectors, selected_neuron_ids
        )
        self._set_vector_data(
            active_layer, active_vectors, active_ids, active=True
        )
        if box_label_layer is not None:
            self._set_box_label_data(
                box_label_layer,
                label_points,
                label_neuron_ids,
                display_texts,
            )

    def _set_vector_data(
        self,
        layer: Vectors,
        vector_blocks: list[np.ndarray],
        neuron_ids: list[int],
        *,
        active: bool = False,
    ) -> None:
        ndim = layer.ndim
        data = (
            np.concatenate(vector_blocks, axis=0)
            if vector_blocks
            else np.empty((0, 2, ndim), dtype=float)
        )
        layer.data = data
        layer.features = {"neuron_id": np.asarray(neuron_ids, dtype=int)}
        if active:
            layer.edge_color = "yellow"
        elif neuron_ids:
            layer.edge_color = np.asarray(
                [self._display_color(item_id) for item_id in neuron_ids]
            )
        else:
            layer.edge_color = "white"

    def _set_box_label_data(
        self,
        layer: Points,
        points: list[np.ndarray],
        neuron_ids: list[int],
        display_texts: list[str],
    ) -> None:
        ndim = layer.ndim
        layer.data = (
            np.asarray(points, dtype=float).reshape(-1, ndim)
            if points
            else np.empty((0, ndim), dtype=float)
        )
        layer.features = {
            "neuron_id": np.asarray(neuron_ids, dtype=int),
            "display_text": np.asarray(display_texts, dtype=str),
        }
        layer.editable = False
        layer.visible = self.show_box_labels_checkbox.isChecked()

    def _box_label_text(self, neuron_id: int, biological: str) -> str:
        biological = biological.strip()
        mode = self.box_label_mode_combo.currentData()
        if mode == LABEL_MODE_DIGITAL:
            return str(neuron_id)
        if mode == LABEL_MODE_DIGITAL_BIOLOGICAL and biological:
            return f"{neuron_id} · {biological}"
        return biological or str(neuron_id)

    def _locate_active_box(self) -> None:
        if (
            self.roi_dataset is None
            or self.current_image is None
            or self.active_id is None
        ):
            return
        volume_index = self._current_volume_index()
        box = (
            self.proofread_store.resolve(volume_index, self.active_id)
            if self.proofread_store is not None and volume_index is not None
            else self.roi_dataset.get_box(self._viewer_time(), self.active_id)
        )
        if box is None:
            self.update_status(
                f"Neuron {self.active_id} is missing at this volume", "orange"
            )
            return
        if not self._id_in_active_z_layer(self.active_id):
            self.update_status(
                f"Neuron {self.active_id} is outside the current Z layer",
                "orange",
            )
            return

        z, y, x = box.center_zyx
        steps = list(self.viewer.dims.current_step)
        if self.current_image.ndim == 4 and len(steps) >= 4:
            z_axis = len(steps) - 3
        else:
            z_axis = len(steps) - 3
        z_size = self._shape_zyx()[0]
        steps[z_axis] = int(np.clip(round(z), 0, z_size - 1))
        self.viewer.dims.current_step = tuple(steps)

        point = (
            (self._viewer_time(), z, y, x)
            if self.current_image.ndim == 4
            else (z, y, x)
        )
        world = self.current_image.data_to_world(point)
        self.viewer.camera.center = tuple(world[-3:])

    # ------------------------------------------------------------------
    # Annotation and Excel
    # ------------------------------------------------------------------
    def _annotation_names(self) -> dict[int, str]:
        names: dict[int, str] = {}
        for row in range(self.annotation_table.rowCount()):
            digital = self.annotation_table.item(row, 0)
            biological = self.annotation_table.item(row, 1)
            if digital is None:
                continue
            try:
                item_id = int(digital.text())
            except ValueError:
                continue
            names[item_id] = (
                biological.text() if biological is not None else ""
            )
        return names

    def _annotation_rows(
        self,
    ) -> dict[int, tuple[str, str]]:
        rows: dict[int, tuple[str, str]] = {}
        for row in range(self.annotation_table.rowCount()):
            digital = self.annotation_table.item(row, 0)
            if digital is None:
                continue
            try:
                item_id = int(digital.text())
            except ValueError:
                continue
            biological = self.annotation_table.item(row, 1)
            annotation = self.annotation_table.item(row, 2)
            rows[item_id] = (
                biological.text() if biological is not None else "",
                annotation.text() if annotation is not None else "",
            )
        return rows

    def _set_annotation_rows(
        self,
        rows: list[tuple[int, str, str]],
    ) -> None:
        self._ui_sync = True
        self.annotation_table.blockSignals(True)
        try:
            self.annotation_table.setRowCount(len(rows))
            for row, (item_id, biological, annotation) in enumerate(rows):
                digital_item = QTableWidgetItem(str(item_id))
                digital_item.setTextAlignment(Qt.AlignCenter)
                self.annotation_table.setItem(row, 0, digital_item)
                self.annotation_table.setItem(
                    row, 1, QTableWidgetItem(biological)
                )
                self.annotation_table.setItem(
                    row, 2, QTableWidgetItem(annotation)
                )
        finally:
            self.annotation_table.blockSignals(False)
            self._ui_sync = False
        self._recompute_search_matches()
        self._refresh_roi_layers()

    def _sync_annotation_ids(self) -> int:
        existing = self._annotation_rows()
        rows = [
            (item_id, *existing.get(item_id, ("", "")))
            for item_id in self._available_ids
        ]
        self._set_annotation_rows(rows)
        return len(rows)

    def load_current_ids_to_annotation(self) -> None:
        count = self._sync_annotation_ids()
        self.update_status(
            f"Loaded {count} identities into annotation table", "green"
        )

    def _sync_annotation_to_active(self) -> None:
        if self.active_id is None:
            self._ui_sync = True
            try:
                self.annotation_table.clearSelection()
            finally:
                self._ui_sync = False
            return
        self._ui_sync = True
        try:
            for row in range(self.annotation_table.rowCount()):
                item = self.annotation_table.item(row, 0)
                if item is not None and item.text() == str(self.active_id):
                    self.annotation_table.selectRow(row)
                    self.annotation_table.scrollToItem(item)
                    break
        finally:
            self._ui_sync = False

    def _on_annotation_selection_changed(self) -> None:
        if self._ui_sync:
            return
        rows = self.annotation_table.selectionModel().selectedRows()
        if not rows:
            return
        item = self.annotation_table.item(rows[0].row(), 0)
        if item is None:
            return
        try:
            item_id = int(item.text())
        except ValueError:
            return
        if item_id not in self._available_ids:
            return
        self.activate_id(item_id, locate=True)

    def _on_annotation_item_changed(self, item: QTableWidgetItem) -> None:
        if self._ui_sync:
            return
        if item.column() in (0, 1):
            self._recompute_search_matches()
            self._refresh_roi_layers()

    def save_annotation_to_excel(self) -> None:
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(
                self, "Excel unavailable", "Install the 'excel' extra first."
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save neuron annotations",
            "neuron_annotations.xlsx",
            "Excel workbooks (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        sheet_name = (
            self.sheet_name_input.text().strip() or "Neuron Annotations"
        )

        try:
            if Path(path).exists():
                workbook = load_workbook(path)
                if sheet_name in workbook.sheetnames:
                    reply = QMessageBox.question(
                        self,
                        "Sheet exists",
                        f"Overwrite sheet '{sheet_name}'?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No,
                    )
                    if reply != QMessageBox.Yes:
                        workbook.close()
                        return
                    workbook.remove(workbook[sheet_name])
                sheet = workbook.create_sheet(sheet_name)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = sheet_name

            sheet.append(["digital", "biological", "annotation"])
            for cell in sheet[1]:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            for item_id, (
                biological,
                annotation,
            ) in self._annotation_rows().items():
                sheet.append([item_id, biological, annotation])
            for column in sheet.columns:
                width = max(
                    (
                        len(str(cell.value))
                        for cell in column
                        if cell.value is not None
                    ),
                    default=0,
                )
                sheet.column_dimensions[column[0].column_letter].width = (
                    width + 2
                )
            workbook.save(path)
            workbook.close()
        except (OSError, PermissionError, ValueError) as error:
            QMessageBox.critical(
                self, "Save failed", f"Could not save workbook:\n{error}"
            )
            self.update_status("Annotation save failed", "red")
            return
        self.update_status(f"Saved annotations to {sheet_name}", "green")

    def load_excel_to_annotation(self) -> None:
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(
                self, "Excel unavailable", "Install the 'excel' extra first."
            )
            return
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Load neuron annotations",
            "",
            "Excel workbooks (*.xlsx)",
        )
        if not path:
            return
        requested = self.sheet_name_input.text().strip()

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            if requested and requested in workbook.sheetnames:
                sheet = workbook[requested]
            else:
                sheet = workbook[workbook.sheetnames[0]]
                self.sheet_name_input.setText(sheet.title)

            rows: list[tuple[int, str, str]] = []
            for values in sheet.iter_rows(values_only=True):
                if not values or values[0] is None:
                    continue
                try:
                    numeric_id = float(values[0])
                except (TypeError, ValueError):
                    continue
                if (
                    not np.isfinite(numeric_id)
                    or not numeric_id.is_integer()
                    or numeric_id < 0
                ):
                    continue
                item_id = int(numeric_id)
                biological = (
                    ""
                    if len(values) < 2 or values[1] is None
                    else str(values[1])
                )
                annotation = (
                    ""
                    if len(values) < 3 or values[2] is None
                    else str(values[2])
                )
                rows.append((item_id, biological, annotation))
            workbook.close()
        except (OSError, PermissionError, ValueError) as error:
            QMessageBox.critical(
                self, "Load failed", f"Could not load workbook:\n{error}"
            )
            self.update_status("Annotation load failed", "red")
            return

        self._set_annotation_rows(rows)
        self.update_status(f"Loaded {len(rows)} annotations", "green")

    # ------------------------------------------------------------------
    # Status helpers
    # ------------------------------------------------------------------
    def _update_info(self) -> None:
        # Keep the compact proofreading status synchronized with all state
        # transitions that refresh the general information panel (including
        # source/Image detach and unload paths).
        self._update_proof_current_box_status()
        mode = "ROI" if self.roi_dataset is not None else "Idle"
        checked = sorted(self.checked_ids)
        text = (
            f"Mode: {mode}\n"
            f"Active: {self.active_id if self.active_id is not None else 'none'}\n"
            f"Checked: {checked[:12]}"
        )
        store = self.proofread_store
        if store is not None:
            proof_mode = "On" if self.proofreading_enabled else "Off"
            if self._proof_detached:
                proof_mode += " (paused)"
            text += (
                f"\nProof: {proof_mode}; "
                f"dirty={'yes' if store.dirty else 'no'}; "
                f"modified={len(store.modified_observations)} observations/"
                f"{len(store.modified_ids)} neurons"
            )
            if self._proof_size_draft_dirty:
                text += "; size draft"
        active_range = self._active_z_range()
        if self._z_ranges:
            z_view = (
                "All"
                if active_range is None
                else f"Layer {active_range.index + 1} "
                f"[{active_range.start},{active_range.stop})"
            )
            text += f"\nZ view: {z_view}"
        if len(checked) > 12:
            text += f" … (+{len(checked) - 12})"
        self.info_text.setText(text)

    def update_status(self, message: str, color: str = "black") -> None:
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {color};")

LabelManager = NeuronAnnotatorWidget

__all__ = (
    "EXCEL_AVAILABLE",
    "LabelManager",
    "NeuronAnnotatorWidget",
    "PROOF_TARGET_EDGE_WIDTH",
    "PROOF_TARGET_HALF_LENGTH",
    "ROLE_ACTIVE",
    "ROLE_BOX_LABELS",
    "ROLE_KEY",
    "ROLE_PROOF_TARGET",
    "ROLE_SELECTED",
    "ROLE_Z_IMAGE",
)
